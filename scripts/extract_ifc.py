#!/usr/bin/env python3
import ifcopenshell
import ifcopenshell.geom
import ifcopenshell.util.element
import json, re, os, base64, tempfile, io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

import sys as _sys

# デフォルトパス（CLI引数で上書き可能）
IFC_PATH = ""
OUTPUT_PATH = ""
KIT_OUTPUT_PATH = ""  # キット積算Excel出力パス（空なら自動生成）

TYPE_NAMES = {
    "IfcWall": "壁", "IfcWallStandardCase": "壁",
    "IfcSlab": "スラブ", "IfcColumn": "柱", "IfcBeam": "梁",
    "IfcDoor": "ドア", "IfcWindow": "窓", "IfcStair": "階段",
    "IfcRailing": "手摺", "IfcRoof": "屋根",
    "IfcMember": "部材",
    "IfcPlate": "板", "IfcCurtainWall": "カーテンウォール",
    "IfcFooting": "基礎",
}

# 除外するIFCタイプ（積算不要）
SKIP_TYPES = {"IfcBuildingElementProxy", "IfcCovering"}

# 壁のサブ分類（要素名ベース）
WALL_SUB = {
    "log": "ログ壁",
    "majikiri": "間仕切壁",
    "kiso": "基礎",
    "dodai": "土台",
}

# 除外する壁の要素名プレフィクス（断熱壁など積算不要）
SKIP_WALL_PREFIX = {"I"}

def classify_wall(elem_name):
    """壁の要素名からサブ分類を返す"""
    if not elem_name:
        return "壁（その他）"
    for key, label in WALL_SUB.items():
        if elem_name.startswith(key):
            return label
    return "壁（その他）"

# スラブのサブ分類（要素名 + レイヤーベース）
def classify_slab(elem_name, layer_name):
    """スラブの要素名・レイヤーからサブ分類を返す"""
    n = (elem_name or "").lower()
    l = (layer_name or "").lower()
    if "yane" in n or "屋根" in l:
        return "屋根"
    if "terrace" in n:
        return "テラス"
    if "balcony" in n or "balconi" in n:
        return "バルコニー"
    if "2-yuka" in n or "2f" in n:
        return "2F床"
    if "yuka" in n or "1-yuka" in n:
        return "1F床"
    return "スラブ（その他）"

def get_majikiri_panel_info(element):
    """間仕切壁のパネル枚数と芯材(Stud)情報を取得。
    材質レイヤーから: パネル(20mm)=面材, 胴縁(95mm)=Stud
    フォールバック: 壁の厚み(Width)で判定 → 135mm=両面(2枚), 115mm=片面(1枚)
    Returns: (panel_count, stud_thickness_mm)
    """
    panel_count = 0
    stud_thick = 0
    detected_by_material = False
    try:
        material = ifcopenshell.util.element.get_material(element)
        if material:
            layers = []
            if material.is_a("IfcMaterialLayerSetUsage"):
                layers = material.ForLayerSet.MaterialLayers
            elif material.is_a("IfcMaterialLayerSet"):
                layers = material.MaterialLayers
            if layers:
                for layer in layers:
                    if not layer.Material:
                        continue
                    name = layer.Material.Name or ""
                    thick = layer.LayerThickness or 0
                    if "パネル" in name or "panel" in name.lower():
                        panel_count += 1
                    elif "胴縁" in name or "stud" in name.lower():
                        stud_thick = thick
                detected_by_material = True
            elif material.is_a("IfcMaterialList"):
                # IfcMaterialListの場合は名前でカウント
                for m in material.Materials:
                    if m and m.Name:
                        if "パネル" in m.Name or "panel" in m.Name.lower():
                            panel_count += 1
                        elif "胴縁" in m.Name or "stud" in m.Name.lower():
                            stud_thick = 95  # default
                detected_by_material = True
    except:
        pass

    # 材質から判定できなかった場合 → 壁の厚み(Width)でフォールバック
    # 135mm = パネル(20)+胴縁(95)+パネル(20) → 両面(2枚)
    # 115mm = 胴縁(95)+パネル(20)            → 片面(1枚)
    if not detected_by_material or panel_count == 0:
        try:
            psets = ifcopenshell.util.element.get_psets(element)
            bq = psets.get('BaseQuantities', {})
            aq = psets.get('ArchiCADQuantities', {})
            width = bq.get('Width') or aq.get('幅') or aq.get('厚さ')
            if isinstance(width, (int, float)) and width > 0:
                w_mm = width if width > 10 else width * 1000  # m→mm変換
                if abs(w_mm - 135) < 5:
                    panel_count = 2
                    stud_thick = 95
                elif abs(w_mm - 115) < 5:
                    panel_count = 1
                    stud_thick = 95
        except:
            pass

    return panel_count, stud_thick

def get_material_info(element):
    try:
        material = ifcopenshell.util.element.get_material(element)
        if material:
            if material.is_a("IfcMaterial"):
                return material.Name or ""
            elif material.is_a("IfcMaterialLayerSetUsage"):
                return " / ".join(l.Material.Name for l in material.ForLayerSet.MaterialLayers if l.Material)
            elif material.is_a("IfcMaterialLayerSet"):
                return " / ".join(l.Material.Name for l in material.MaterialLayers if l.Material)
            elif material.is_a("IfcMaterialList"):
                return " / ".join(m.Name for m in material.Materials if m)
    except:
        pass
    return ""

def get_storey(element):
    try:
        container = ifcopenshell.util.element.get_container(element)
        if container and container.is_a("IfcBuildingStorey"):
            return container.Name or ""
    except:
        pass
    return ""

def extract_key_props(psets, ifc_type):
    """重要なプロパティを抽出"""
    arch = psets.get('ArchiCADProperties', {})
    bq = psets.get('BaseQuantities', {})
    aq = psets.get('ArchiCADQuantities', {})

    marker = arch.get('マーカーテキスト', '')
    nominal = arch.get('公称幅x高さ', '')
    nominal_thick = arch.get('公称幅x高さx厚さ', '')
    wall_struct = arch.get('壁構造', '')
    layer = arch.get('レイヤー', '')
    parent_id = arch.get('親ID', '')

    # AC_Pset_* から建具情報
    maker = ''
    model_no = ''
    kind = ''
    has_gakubuchi = ''
    has_kiryoke = ''
    gakubuchi_w = ''
    gakubuchi_t = ''
    for pn, pp in psets.items():
        if pn.startswith('AC_Pset_') and pn != 'AC_Pset_RenovationAndPhasing':
            maker = pp.get('メーカー', pp.get('Maker', maker))
            model_no = pp.get('型番', pp.get('Model', model_no))
            kind = pp.get('建具種類', kind)
            if '額縁の有無' in pp:
                has_gakubuchi = 'あり' if pp['額縁の有無'] else 'なし'
            if '霧除けの有無' in pp:
                has_kiryoke = 'あり' if pp['霧除けの有無'] else 'なし'
    for pn, pp in psets.items():
        if pn.startswith('AC_Equantity_'):
            gakubuchi_w = pp.get('額縁幅', '')
            gakubuchi_t = pp.get('額縁の厚み', '')

    # BaseQuantities → ArchiCADQuantities フォールバック
    # 壁: BQに GrossSideArea/NetSideArea あり、Area なし
    # 柱/Proxy: BQ自体なし、AQのみ
    # ドア/窓: BQに Area あり
    def pick(bq_keys, aq_keys):
        for k in bq_keys:
            v = bq.get(k)
            if v is not None and v != '':
                return v
        for k in aq_keys:
            v = aq.get(k)
            if v is not None and v != '':
                return v
        return ''

    bq_width = pick(['Width'], ['幅', '厚さ'])
    bq_height = pick(['Height'], ['高さ'])
    bq_depth = pick(['Depth', 'Thickness'], [])
    bq_length = pick(['Length', 'NetLength'], ['長さ(A)', '基準線長さ', '3D長さ'])
    bq_perimeter = pick(['Perimeter'], ['平面図外周'])

    # 面積: 壁はNetSideArea(開口除く壁面面積)、ドア/窓はArea、柱/ProxyはAQ表面積
    bq_area = pick(
        ['Area', 'GrossArea', 'NetArea', 'NetSideArea', 'GrossSideArea'],
        ['表面積'])
    bq_volume = pick(
        ['Volume', 'NetVolume', 'GrossVolume'],
        ['正味体積'])

    gl_height = aq.get('GLからの高度', '')
    # 壁の開口面積（窓+ドア）
    aq_window_area = aq.get('窓面積', '')
    aq_door_area = aq.get('ドア面積', '')
    opening_area = ''
    if isinstance(aq_window_area, (int, float)) and isinstance(aq_door_area, (int, float)):
        opening_area = aq_window_area + aq_door_area
    elif isinstance(aq_window_area, (int, float)):
        opening_area = aq_window_area
    elif isinstance(aq_door_area, (int, float)):
        opening_area = aq_door_area

    return {
        'マーカー': marker,
        'メーカー': maker,
        '型番': model_no,
        '建具種類': kind,
        'レイヤー': layer,
        '親ID': parent_id,
        '壁構造': wall_struct,
        '公称寸法': nominal_thick or nominal,
        '幅(mm)': bq_width,
        '高さ(mm)': bq_height,
        '厚さ/奥行(mm)': bq_depth,
        '長さ(mm)': bq_length,
        '面積(m²)': bq_area,
        '体積(m³)': bq_volume,
        '周長(mm)': bq_perimeter,
        'GL高度(mm)': gl_height,
        '開口面積(m²)': opening_area,
        '額縁': has_gakubuchi,
        '霧除け': has_kiryoke,
        '額縁幅(mm)': gakubuchi_w,
        '額縁厚(mm)': gakubuchi_t,
    }

def main():
    global IFC_PATH, OUTPUT_PATH, KIT_OUTPUT_PATH

    # CLI: python extract_ifc.py <input.ifc> <output.xlsx> [kit_output.xlsx]
    if len(_sys.argv) >= 3:
        IFC_PATH = _sys.argv[1]
        OUTPUT_PATH = _sys.argv[2]
        if len(_sys.argv) >= 4:
            KIT_OUTPUT_PATH = _sys.argv[3]
    elif not IFC_PATH or not OUTPUT_PATH:
        print("Usage: python extract_ifc.py <input.ifc> <output.xlsx> [kit_output.xlsx]")
        _sys.exit(1)

    print(f"IFC読み込み中... {IFC_PATH}")
    ifc_file = ifcopenshell.open(IFC_PATH)
    target_types = [t for t in TYPE_NAMES.keys() if t not in SKIP_TYPES]
    all_elements = []
    seen_ids = set()  # IfcWall/IfcWallStandardCase 重複除去用

    for ifc_type in target_types:
        elements = ifc_file.by_type(ifc_type)
        if not elements:
            continue
        added = 0
        for elem in elements:
            # IfcWallStandardCaseはIfcWallのサブタイプ → 重複除去
            gid = elem.GlobalId
            if gid in seen_ids:
                continue
            seen_ids.add(gid)
            added += 1

            type_name_jp = TYPE_NAMES.get(ifc_type, ifc_type)

            etype = ifcopenshell.util.element.get_type(elem)
            elem_type_name = etype.Name if etype else ""
            material = get_material_info(elem)
            storey = get_storey(elem)
            psets = ifcopenshell.util.element.get_psets(elem)
            kp = extract_key_props(psets, ifc_type)

            # 壁はサブ分類（ログ壁/間仕切壁/基礎/土台）
            if type_name_jp == "壁":
                # 除外対象の壁プレフィクスをスキップ
                ename = elem.Name or ""
                if any(ename.startswith(p) for p in SKIP_WALL_PREFIX):
                    continue
                type_name_jp = classify_wall(ename)

            # スラブはサブ分類（1F床/2F床/テラス/バルコニー/屋根）
            if type_name_jp == "スラブ":
                type_name_jp = classify_slab(elem.Name, kp['レイヤー'])

            # 間仕切壁のパネル情報
            panel_count = ''
            panel_area = ''
            stud_area = ''
            if type_name_jp == "間仕切壁":
                pc, st = get_majikiri_panel_info(elem)
                panel_count = pc if pc > 0 else 0
                # GrossSideArea = 片面の壁面面積
                area_val = kp['面積(m²)']
                if isinstance(area_val, (int, float)) and area_val > 0:
                    stud_area = area_val  # Stud面積 = 片面面積
                    panel_area = area_val * pc  # パネル面積 = 片面面積 × パネル枚数

            row = {
                "部材分類": type_name_jp,
                "要素名": elem.Name or "",
                "マーカー": kp['マーカー'],
                "型式名": elem_type_name,
                "階": storey,
                "材質": material,
                "メーカー": kp['メーカー'],
                "型番": kp['型番'],
                "建具種類": kp['建具種類'],
                "レイヤー": kp['レイヤー'],
                "壁構造": kp['壁構造'],
                "公称寸法": kp['公称寸法'],
                "幅(mm)": kp['幅(mm)'],
                "高さ(mm)": kp['高さ(mm)'],
                "厚さ/奥行(mm)": kp['厚さ/奥行(mm)'],
                "長さ(mm)": kp['長さ(mm)'],
                "面積(m²)": kp['面積(m²)'],
                "体積(m³)": kp['体積(m³)'],
                "パネル枚数": panel_count,
                "パネル面積(m²)": panel_area,
                "Stud面積(m²)": stud_area,
                "周長(mm)": kp['周長(mm)'],
                "GL高度(mm)": kp['GL高度(mm)'],
                "開口面積(m²)": kp['開口面積(m²)'],
                "額縁": kp['額縁'],
                "霧除け": kp['霧除け'],
                "額縁幅(mm)": kp['額縁幅(mm)'],
                "額縁厚(mm)": kp['額縁厚(mm)'],
                "親ID": kp['親ID'],
                "GlobalId": elem.GlobalId,
            }
            all_elements.append(row)
        if added:
            print(f"  {ifc_type}: {added}")

    print(f"\n合計 {len(all_elements)} 部材（重複除去済み）")

    # === Excel出力 ===
    wb = Workbook()
    hfont = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cfont = Font(name="Arial", size=10)
    border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    num_right = Alignment(horizontal="right")

    # ---------- Sheet 1: 部材一覧 ----------
    ws = wb.active
    ws.title = "部材一覧"
    headers = list(all_elements[0].keys())
    num_cols = {"幅(mm)", "高さ(mm)", "厚さ/奥行(mm)", "長さ(mm)", "面積(m²)", "体積(m³)",
                "パネル枚数", "パネル面積(m²)", "Stud面積(m²)",
                "周長(mm)", "GL高度(mm)", "開口面積(m²)", "額縁幅(mm)", "額縁厚(mm)"}

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

    for ri, elem in enumerate(all_elements, 2):
        for ci, key in enumerate(headers, 1):
            val = elem.get(key, "")
            if val == '' or val is None:
                val = None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cfont; c.border = border
            if ri % 2 == 0:
                c.fill = alt_fill
            if key in num_cols and isinstance(val, (int, float)):
                c.alignment = num_right
                if "m²" in key:
                    c.number_format = '#,##0.0000'
                elif "m³" in key:
                    c.number_format = '#,##0.000000'
                else:
                    c.number_format = '#,##0.0'

    col_w = {"部材分類":12, "要素名":14, "マーカー":16, "型式名":22, "階":12, "材質":25,
             "メーカー":10, "型番":10, "建具種類":8, "レイヤー":14, "壁構造":20, "公称寸法":20,
             "幅(mm)":10, "高さ(mm)":10, "厚さ/奥行(mm)":12, "長さ(mm)":10,
             "面積(m²)":12, "体積(m³)":14,
             "パネル枚数":8, "パネル面積(m²)":14, "Stud面積(m²)":14,
             "周長(mm)":10, "GL高度(mm)":12,
             "開口面積(m²)":12, "額縁":6, "霧除け":6, "額縁幅(mm)":10, "額縁厚(mm)":10,
             "親ID":12, "GlobalId":36}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(h, 12)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_elements)+1}"
    ws.freeze_panes = "A2"

    # ---------- Sheet 2: 建具一覧 ----------
    ws2 = wb.create_sheet("建具一覧")
    fittings = [e for e in all_elements if e["部材分類"] in ("ドア", "窓")]
    f_headers = ["部材分類", "要素名", "マーカー", "建具種類", "メーカー", "型番", "型式名",
                 "幅(mm)", "高さ(mm)", "厚さ/奥行(mm)", "面積(m²)", "体積(m³)",
                 "GL高度(mm)", "開口面積(m²)", "額縁", "霧除け", "額縁幅(mm)", "額縁厚(mm)",
                 "壁構造", "親ID", "階"]

    for ci, h in enumerate(f_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

    # マーカーでソート
    fittings.sort(key=lambda x: (x["部材分類"], x.get("マーカー") or "zzz"))
    for ri, elem in enumerate(fittings, 2):
        for ci, key in enumerate(f_headers, 1):
            val = elem.get(key)
            if val == '' or val is None:
                val = None
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = cfont; c.border = border
            if ri % 2 == 0:
                c.fill = alt_fill
            if key in num_cols and isinstance(val, (int, float)):
                c.alignment = num_right
                if "m²" in key:
                    c.number_format = '#,##0.0000'
                elif "m³" in key:
                    c.number_format = '#,##0.000000'
                else:
                    c.number_format = '#,##0.0'

    f_col_w = {"部材分類":8, "要素名":8, "マーカー":16, "建具種類":8, "メーカー":10, "型番":10,
               "型式名":22, "幅(mm)":10, "高さ(mm)":10, "厚さ/奥行(mm)":12, "面積(m²)":12,
               "体積(m³)":14, "GL高度(mm)":12, "開口面積(m²)":12, "額縁":6, "霧除け":6,
               "額縁幅(mm)":10, "額縁厚(mm)":10, "壁構造":18, "親ID":12, "階":12}
    for ci, h in enumerate(f_headers, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = f_col_w.get(h, 10)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(f_headers))}{len(fittings)+1}"
    ws2.freeze_panes = "A2"

    # ---------- Sheet 3: 分類別集計 ----------
    # 集計ルール:
    #   ログ壁        → 面積(m²)
    #   間仕切壁      → パネル面積 / Stud面積
    #   基礎          → 長さ(m)
    #   土台          → 長さ(m)
    #   屋根(yane1)   → 型式名ごとに集計。破風・鼻隠しは長さ(m)、屋根本体は面積(m²)
    #   床            → 要素名・型式名ごとに面積(m²)
    #   柱(チャネル柱) → 木口カバー総長さ(m)
    #   梁            → 体積(m³)
    #   ドア・窓      → 集計不要
    #   手摺・階段    → 数量のみ
    ws3 = wb.create_sheet("分類別集計")

    s_headers = ["分類", "詳細", "数量", "集計値", "単位", "備考"]
    for ci, h in enumerate(s_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

    summary_rows = []  # (分類, 詳細, 数量, 値, 単位, 備考)

    # --- ログ壁 ---
    log_elems = [e for e in all_elements if e["部材分類"] == "ログ壁"]
    if log_elems:
        total_area = sum(e["面積(m²)"] for e in log_elems if isinstance(e.get("面積(m²)"), (int, float)))
        summary_rows.append(("ログ壁", "", len(log_elems), round(total_area, 2), "m²", ""))

    # --- 間仕切壁 ---
    maji_elems = [e for e in all_elements if e["部材分類"] == "間仕切壁"]
    if maji_elems:
        total_panel = sum(e["パネル面積(m²)"] for e in maji_elems if isinstance(e.get("パネル面積(m²)"), (int, float)))
        total_stud = sum(e["Stud面積(m²)"] for e in maji_elems if isinstance(e.get("Stud面積(m²)"), (int, float)))
        summary_rows.append(("間仕切壁", "パネル面積", len(maji_elems), round(total_panel, 2), "m²", ""))
        summary_rows.append(("", "Stud面積", "", round(total_stud, 2), "m²", ""))

    # --- 基礎 → 長さ ---
    kiso_elems = [e for e in all_elements if e["部材分類"] == "基礎"]
    if kiso_elems:
        total_len = sum(e["長さ(mm)"] for e in kiso_elems if isinstance(e.get("長さ(mm)"), (int, float)))
        summary_rows.append(("基礎", "", len(kiso_elems), round(total_len / 1000, 2), "m", ""))

    # --- 土台 → 長さ ---
    dodai_elems = [e for e in all_elements if e["部材分類"] == "土台"]
    if dodai_elems:
        total_len = sum(e["長さ(mm)"] for e in dodai_elems if isinstance(e.get("長さ(mm)"), (int, float)))
        summary_rows.append(("土台", "", len(dodai_elems), round(total_len / 1000, 2), "m", ""))

    # --- 屋根 → 型式名ごと。破風・鼻隠しは長さ、屋根本体は面積 ---
    yane_elems = [e for e in all_elements if e["部材分類"] == "屋根"]
    if yane_elems:
        # 型式名でグループ化
        yane_by_type = defaultdict(list)
        for e in yane_elems:
            tn = e.get("型式名") or "（不明）"
            yane_by_type[tn].append(e)

        for tn, elems in sorted(yane_by_type.items()):
            is_hafu = "破風" in tn or "鼻隠" in tn
            if is_hafu:
                # 長さ = GrossArea / (Width/1000)
                total_len = 0
                for e in elems:
                    area = e.get("面積(m²)")
                    width = e.get("幅(mm)")
                    if isinstance(area, (int, float)) and isinstance(width, (int, float)) and width > 0:
                        total_len += area / (width / 1000)
                summary_rows.append(("屋根", tn, len(elems), round(total_len, 2), "m", "長さ"))
            else:
                total_area = sum(e["面積(m²)"] for e in elems if isinstance(e.get("面積(m²)"), (int, float)))
                summary_rows.append(("屋根", tn, len(elems), round(total_area, 2), "m²", "面積"))

    # --- 床 → 要素名・型式名ごとに面積 ---
    floor_cats = {"1F床", "2F床", "テラス", "バルコニー"}
    floor_elems = [e for e in all_elements if e["部材分類"] in floor_cats]
    if floor_elems:
        # (要素名, 型式名) でグループ化
        floor_by_key = defaultdict(list)
        for e in floor_elems:
            key = (e.get("要素名") or "", e.get("型式名") or "")
            floor_by_key[key].append(e)

        for (ename, tname), elems in sorted(floor_by_key.items()):
            total_area = sum(e["面積(m²)"] for e in elems if isinstance(e.get("面積(m²)"), (int, float)))
            cat = elems[0]["部材分類"]
            summary_rows.append(("床", f"{ename} / {tname}", len(elems), round(total_area, 2), "m²", cat))

    # --- 柱(チャネル柱) → 木口カバー総長さ ---
    col_elems = [e for e in all_elements if e["部材分類"] == "柱"]
    if col_elems:
        total_h = sum(e["高さ(mm)"] for e in col_elems if isinstance(e.get("高さ(mm)"), (int, float)))
        summary_rows.append(("木口カバー", "チャネル柱", len(col_elems), round(total_h / 1000, 2), "m", "総長さ"))

    # --- 梁 → m³ ---
    beam_elems = [e for e in all_elements if e["部材分類"] == "梁"]
    if beam_elems:
        total_vol = sum(e["体積(m³)"] for e in beam_elems if isinstance(e.get("体積(m³)"), (int, float)))
        summary_rows.append(("梁", "", len(beam_elems), round(total_vol, 4), "m³", ""))

    # --- 手摺・階段 → 数量のみ ---
    for cat in ["手摺", "階段"]:
        cat_elems = [e for e in all_elements if e["部材分類"] == cat]
        if cat_elems:
            summary_rows.append((cat, "", len(cat_elems), "", "", ""))

    # --- その他（スラブその他等） ---
    shown_cats = {"ログ壁", "間仕切壁", "基礎", "土台", "屋根",
                  "1F床", "2F床", "テラス", "バルコニー",
                  "柱", "梁", "ドア", "窓", "手摺", "階段"}
    for e in all_elements:
        cat = e["部材分類"]
        if cat not in shown_cats:
            shown_cats.add(cat)
            cat_elems = [x for x in all_elements if x["部材分類"] == cat]
            total_area = sum(x["面積(m²)"] for x in cat_elems if isinstance(x.get("面積(m²)"), (int, float)))
            summary_rows.append((cat, "", len(cat_elems), round(total_area, 2) if total_area else "", "m²" if total_area else "", ""))

    # 書き出し
    bf = Font(name="Arial", bold=True, size=10)
    cat_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    prev_cat = None
    for ri, (cat, detail, count, val, unit, note) in enumerate(summary_rows, 2):
        is_cat_row = (cat != "" and cat != prev_cat)
        ws3.cell(row=ri, column=1, value=cat).font = bf if is_cat_row else cfont
        ws3.cell(row=ri, column=2, value=detail).font = cfont
        ws3.cell(row=ri, column=3, value=count if count != "" else None).font = cfont
        c4 = ws3.cell(row=ri, column=4, value=val if val != "" else None)
        c4.font = cfont
        if isinstance(val, (int, float)):
            if unit == "m³":
                c4.number_format = '#,##0.0000'
            else:
                c4.number_format = '#,##0.00'
        ws3.cell(row=ri, column=5, value=unit).font = cfont
        ws3.cell(row=ri, column=6, value=note).font = cfont
        for ci in range(1, 7):
            ws3.cell(row=ri, column=ci).border = border
            if is_cat_row:
                ws3.cell(row=ri, column=ci).fill = cat_fill
        if cat:
            prev_cat = cat

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 36
    ws3.column_dimensions["C"].width = 8
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 6
    ws3.column_dimensions["F"].width = 10

    # ---------- Sheet 4: マーカー別集計 ----------
    ws4 = wb.create_sheet("マーカー別集計")
    marker_stats = defaultdict(lambda: {"count": 0, "category": "", "maker": "", "model": "",
                                         "sizes": set()})
    for e in all_elements:
        m = e.get("マーカー")
        if not m:
            continue
        marker_stats[m]["count"] += 1
        marker_stats[m]["category"] = e["部材分類"]
        if e.get("メーカー"):
            marker_stats[m]["maker"] = e["メーカー"]
        if e.get("型番"):
            marker_stats[m]["model"] = e["型番"]
        w = e.get("幅(mm)")
        h = e.get("高さ(mm)")
        if w and h:
            marker_stats[m]["sizes"].add(f"{w}×{h}")

    m_headers = ["マーカー", "部材分類", "数量", "メーカー", "型番", "サイズ(mm)"]
    for ci, h in enumerate(m_headers, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

    sorted_markers = sorted(marker_stats.items(), key=lambda x: (x[1]["category"], x[0]))
    for ri, (mname, stats) in enumerate(sorted_markers, 2):
        ws4.cell(row=ri, column=1, value=mname).font = cfont
        ws4.cell(row=ri, column=2, value=stats["category"]).font = cfont
        ws4.cell(row=ri, column=3, value=stats["count"]).font = cfont
        ws4.cell(row=ri, column=4, value=stats["maker"]).font = cfont
        ws4.cell(row=ri, column=5, value=stats["model"]).font = cfont
        ws4.cell(row=ri, column=6, value=", ".join(sorted(stats["sizes"]))).font = cfont
        for ci in range(1, 7):
            ws4.cell(row=ri, column=ci).border = border
            if ri % 2 == 0:
                ws4.cell(row=ri, column=ci).fill = alt_fill

    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 8
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 10
    ws4.column_dimensions["F"].width = 30
    ws4.auto_filter.ref = f"A1:F{len(sorted_markers)+1}"
    ws4.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Excel保存: {OUTPUT_PATH}")

    # サマリ表示
    print("\n=== 分類別集計 ===")
    for cat, detail, count, val, unit, note in summary_rows:
        count_s = f"{count:3d}個" if isinstance(count, int) else "   "
        val_s = f"{val:10.2f}" if isinstance(val, (int, float)) else "          "
        print(f"  {cat:12s} {detail:36s} {count_s}  {val_s} {unit:4s} {note}")

    print("\n=== 建具マーカー ===")
    for mname, stats in sorted_markers:
        sizes = ", ".join(sorted(stats["sizes"]))
        print(f"  {mname:16s} {stats['category']:4s} ×{stats['count']}  {stats['maker']:8s} {sizes}")

    # ---------- キット積算 Excel 自動生成 ----------
    kit_path = generate_kit_estimate(all_elements, marker_stats, summary_rows)
    if kit_path:
        print(f"\nキット積算Excel: {kit_path}")


# ============================================================
# キット積算・価格計算 Excel 自動入力（テンプレート内蔵版）
# ============================================================

# キット積算テンプレート（base64エンコード）
_KIT_TEMPLATE_B64 = ""  # ← _load_kit_template_b64() で遅延読み込み

def _load_kit_template_b64():
    """内蔵テンプレートのbase64文字列を返す（巨大文字列を関数内に閉じ込め）"""
    return (
        'UEsDBBQAAAAIABFNaFxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIABFNaFy5FZzL8gAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNksFOwzAMhl8F5d46bccEUdcL004gITEJxC1KvC2iaaLEqN3bk5atA8ED7Bj7z+fPkmvlhXIBn4PzGMhgvBls20Wh/IodiLwAiOqAVsY8JbrU3LlgJaVn2IOX6kPuEUrOl2CRpJYkYQRmfiayptZKqICSXDjhtZrx/jO0E0wrwBYtdhShyAtgzTjRH4e2hgtghBEGG78LqGfiVP0XO3WAnZJDNHOq7/u8r6Zc2qGAt6fHl2ndzHSRZKcw/YpG0NHjip0nv1YP6+2GNSUvlxmvMn635fdiwUW1eB9df/ldhK3TZmeuzPj2h/FZsKnhz100X1BLAwQUAAAACAARTWhcmVycIxAGAACcJwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWltz2jgUfu+v0Hhn9m0LxjaBtrQTc2l227SZhO1OH4URWI1seWSRhH+/RzYQy5YN7ZJNups8BCzp+85FR+foOHnz7i5i6IaIlPJ4YNkv29a7ty/e4FcyJBFBMBmnr/DACqVMXrVaaQDDOH3JExLD3IKLCEt4FMvWXOBbGi8j1uq0291WhGlsoRhHZGB9XixoQNBUUVpvXyC05R8z+BXLVI1lowETV0EmuYi08vlsxfza3j5lz+k6HTKBbjAbWCB/zm+n5E5aiOFUwsTAamc/VmvH0dJIgILJfZQFukn2o9MVCDINOzqdWM52fPbE7Z+Mytp0NG0a4OPxeDi2y9KLcBwE4FG7nsKd9Gy/pEEJtKNp0GTY9tqukaaqjVNP0/d93+ubaJwKjVtP02t33dOOicat0HgNvvFPh8Ouicar0HTraSYn/a5rpOkWaEJG4+t6EhW15UDTIABYcHbWzNIDll4p+nWUGtkdu91BXPBY7jmJEf7GxQTWadIZljRGcp2QBQ4AN8TRTFB8r0G2iuDCktJckNbPKbVQGgiayIH1R4Ihxdyv/fWXu8mkM3qdfTrOa5R/aasBp+27m8+T/HPo5J+nk9dNQs5wvCwJ8fsjW2GHJ247E3I6HGdCfM/29pGlJTLP7/kK6048Zx9WlrBdz8/knoxyI7vd9lh99k9HbiPXqcCzIteURiRFn8gtuuQROLVJDTITPwidhphqUBwCpAkxlqGG+LTGrBHgE323vgjI342I96tvmj1XoVhJ2oT4EEYa4pxz5nPRbPsHpUbR9lW83KOXWBUBlxjfNKo1LMXWeJXA8a2cPB0TEs2UCwZBhpckJhKpOX5NSBP+K6Xa/pzTQPCULyT6SpGPabMjp3QmzegzGsFGrxt1h2jSPHr+BfmcNQockRsdAmcbs0YhhGm78B6vJI6arcIRK0I+Yhk2GnK1FoG2camEYFoSxtF4TtK0EfxZrDWTPmDI7M2Rdc7WkQ4Rkl43Qj5izouQEb8ehjhKmu2icVgE/Z5ew0nB6ILLZv24fobVM2wsjvdH1BdK5A8mpz/pMjQHo5pZCb2EVmqfqoc0PqgeMgoF8bkePuV6eAo3lsa8UK6CewH/0do3wqv4gsA5fy59z6XvufQ9odK3NyN9Z8HTi1veRm5bxPuuMdrXNC4oY1dyzcjHVK+TKdg5n8Ds/Wg+nvHt+tkkhK+aWS0jFpBLgbNBJLj8i8rwKsQJ6GRbJQnLVNNlN4oSnkIbbulT9UqV1+WvuSi4PFvk6a+hdD4sz/k8X+e0zQszQ7dyS+q2lL61JjhK9LHMcE4eyww7ZzySHbZ3oB01+/ZdduQjpTBTl0O4GkK+A226ndw6OJ6YkbkK01KQb8P56cV4GuI52QS5fZhXbefY0dH758FRsKPvPJYdx4jyoiHuoYaYz8NDh3l7X5hnlcZQNBRtbKwkLEa3YLjX8SwU4GRgLaAHg69RAvJSVWAxW8YDK5CifEyMRehw55dcX+PRkuPbpmW1bq8pdxltIlI5wmmYE2eryt5lscFVHc9VW/Kwvmo9tBVOz/5ZrcifDBFOFgsSSGOUF6ZKovMZU77nK0nEVTi/RTO2EpcYvOPmx3FOU7gSdrYPAjK5uzmpemUxZ6by3y0MCSxbiFkS4k1d7dXnm5yueiJ2+pd3wWDy/XDJRw/lO+df9F1Drn723eP6bpM7SEycecURAXRFAiOVHAYWFzLkUO6SkAYTAc2UyUTwAoJkphyAmPoLvfIMuSkVzq0+OX9FLIOGTl7SJRIUirAMBSEXcuPv75Nqd4zX+iyBbYRUMmTVF8pDicE9M3JD2FQl867aJguF2+JUzbsaviZgS8N6bp0tJ//bXtQ9tBc9RvOjmeAes4dzm3q4wkWs/1jWHvky3zlw2zreA17mEyxDpH7BfYqKgBGrYr66r0/5JZw7tHvxgSCb/NbbpPbd4Ax81KtapWQrET9LB3wfkgZjjFv0NF+PFGKtprGtxtoxDHmAWPMMoWY434dFmhoz1YusOY0Kb0HVQOU/29QNaPYNNByRBV4xmbY2o+ROCjzc/u8NsMLEjuHti78BUEsDBBQAAAAIABFNaFx90mGETiQAAEHtAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1s3V1rcxs3lv0rXO1URlbkER6Nl2S7KrZItid2kvIjM1Vb+4GJaVs1kqmRmNizv34v0I1mP3AvIMnKTm25aIp9AfTtgwM0Di7Q/ejz5uof1x/X6+3sy8X5p+vHex+328vjo6PrXz+uL1bXf9lcrj+B5f3m6mK1hZ9XH46uL6/Wq3ch08X5kWBMH12szj7tPXkUjv109eTR5rft+dmn9U9Xs+vfLi5WV/96uj7ffH68x/figVdnHz5u/YGjJ48uVx/Wr9fbt5eQ/v3Z9s3mJzjQ2o66Mt+dXaw/XZ9tPs2u1u8f7z3lxzWXzqcJSX4+W3++7v09u/64+by8Onv3AhyB62J7M3+tv2w2//Dm5+/8IX/uT+vZl9eX52fBm9m/2j/l3my7uXyxfr99tj4/h9NVe7PVr9uz39c/QY7He79sttvNRbgKuKbtagvH3l9t/mf9KXi0Pl9DYvD1MqSGopqkCVtTkj8Tbm3P0zjQOPSdx/KfAYrvekj5a+v/HSFZhAoEgH9ZXa+fbc7/dvZu+/Hxnt2bvVu/X/12vn21+Vyv20pRvrxfN+fX4f/Z5yYtQPLrb9fgTZsXHLg4+9R8r7609dVPL5AMos0gRhl4hWSQbQY5zoC5VLUZqtIzqDaDKj2DbjPo0gymzWDGGRSSwbYZ7BhWJL1r04cmcdTUXqj609V29eTR1ebz7Cqk9lUsWSylq3RoDL/6FE89sUJCOHr2ybfj19srsJ5Bgdsnb7578eM3/8mFEurEf1e6ar5d+1u1v408mTUJeHPAyvbbwrfkkrnwraULx1X4FkxaX5DQ2vqCpJJSd+lms/+aZhXMSuO/Oa/8t1YCTlE/9ylc8MofcSf//ehoC8j4yzj6FT6ASAeLaGDhGodFBFgEAou/Asm8xxXTVfBcStdAIVm4IuV0uHIjhUm4LYx02kMG3oLXs12JwhjGUiXyqhL+2zHZYKx4C741zW/RYC51i32oNCmFsiG/iOXoUE5VGd4BCL+FcTLUhVS6SWdNB2fCUVlxGxwVrJKoo/O3rxoUQiVaB2X1r1rwpsYjE6AQ1zKEs45azXd7vIrH+fB3uAqfnzd0qQJdnGHihOCCbLhAUEEGKshABX/f80UEyzPUcopa5qhlgVqWqKVOWQaXVzWXJ4geoAplVCjVnZaiqXfF5MmgNzCi7QVi79Bv9GAOVSeEMmZUBQ1+zZlVAj/UMkctC9SyRC11Y9E4fiqPnwplGKKruCFuXPE0Xs2ZbAIvlbiOBi/Mux1uBf6Pe3/tWv/b6zBVyt9l399pyZyLcLEn3IlE7jp1tYO60fm60ZlrK+1UYyfptGb9Xka7cN+CVJVusrXZZXv/i71ZvC82vVXqJuaSVa7RKm8sDr20rl9n/n5wPKufP+SSONUcw2rHlByaUsEF9W73QjPo1FPMaEriLNEkU1c8qHaT7bEN2mOjllPUMkctC9SyRC11yjK4PJtntb1Bj90Ox1RLv8hyMfxd2SFNh8M3AWOc5psx17WGcGcXlYyMd4xkskU7e9QyRy0L1LJELbXNdfauHRdaHHrXEBcbL4+xSgHRFoGPLYdjohngnCjmtC1GFvdtwyFpqgdoi8SJNRnaJUpZtKUovJTxWDjn2LItUhNFysBw4ZwTJUXWbZHErYFZw5rhLDNdbRAjSd+bBbI4Qm2xG95sYVQudqP0F5sPf1udn8dm9l8X4r9TDGvPAjx+8uj9k+eL/eevf3j78un81f7r+XevntX7QJtq73CpHjw4/HNqDE6Nvf/8H8/0rXKpB4+O3j959Dt4+nufyNFb13o7f/Xqx1f7P7/48cfv3/60357M8NAzCSvbk6Tv00OC/vk//vT0+E9vD83h4rsXr+cPDvf20j7MWx8EG/gAQB6Ae4csnWsRc/FENxRtItYBpP7mfHvyzYftyd7eIfw6jKdpTpE+Rx3LkbEcKPkxgys5hD8OvIe3qQs5OduQyryAyjw3tqoEE83dA7QeX0QSL843m6unm9XVu1k8SFCZ96l8m0s1COv4fbPO5lnHh6zbP5UcaPX8h73DHQU5UJADP3AWcoKFfMJCPmAh77GQEyzkIxb2PQVK8khJfuA9vh0lU0ANaSkKaCluREuRoqXI01LcFy3FfdMSqjnLS1HASwG8FCQvBcFLMeGlGPBS9HgpCF4Kmpci8lIceI/vi5eygJcyx0thw51fMM5Nott8tj6DTB98QOK8pOOUd2WoRRgq/w06TjkhqJwQVAJBJUlQSRBUTggqBwSVPYJKgqByQlDZJ6iMBJUH3uP7ImhVQNDqVgQVaYIWdKHVfRG0um+CujxBqwKCVkDQiiRoRRC0mhC0GhC06hG0Igha0QStIkGrA+/xfRFUFRA0N1EJ7rggvcGfoJ20UbvJuxeri7NPq+363dP16mJHTpkmp7orOR1CTnXf5FR5cqq02FHAR4WTURFkVBMyqgEZVY+MiiCjmogdFfmnDryH9yF2dAH3clOfPiTa1E3lZK+OOt0eusVsn6jvSjvOEN7pex9Xsjzx9KRXVJNeUQMLNdkraoKIekJEPSCi7hFRE0TUk15R9XtFHVmpD7zH99UrmgJmmuxtu4p6h9mOka82m/evVu+366ssJ81oMmnBTQ/S5sDjvWZ8IE/2hVQP9m41QQQD/Vtl4w+6SgVvviqnq5bTjwW3f/Q13TIfNsvWVKOoBh1AMzRAWquZtFY3aa0GWqshW2tbSmqqPtrEwKcl7lMdc8iUT5DTxHZpDrxv99UubUG7tGS79ALOr3oJ3ISGo7uWyRezMBXx183Z9TbbOO2dbxgS4YtN84XjfLETvpgJXyzwxZJ8aUuZzvX3eGPTvEF9q2MOmfINctrIG3vgfbwv3rgC3jiSNwLljbgRb9ykU3fjTt19nU79dsECru6vU1d37tRvf023zId16i7dSAXeSN2kkdpJI3XQSB3ZSB3Rqbt040R9qmMOmfIJcrrYON2B9+2eGqcoCN+JXPguhuuEb35d0/Qxu8XV6mKda5aC3bk7R+aV25InTNEoUwSbMEWPmSLYAZRMMSWWQnXngqUZg/pWi3FQrO/bPhTYMgb88z7eF2MKomQiFyXzfZZu+yzec6hhzpv11dXq1zxv7hwl48h8hbj3MBmXWd0oRnGyyD4O7EPv+QtBxMXEJC4mBnEx0YuLCSIuJsZxMSg5co8feA/vYcJCFITBBB0Gg6bS1E9cyMdOest6fpydwkCiI13QjlyMl1S31LtzJEwgUxbi/kNhOk89kaaeAOqhd7SFIEJfYhL6EoPQl+iFvgQR+hLj0BeUHKknDryH90E9mV+dJtqQB9XnSdv0dbKZsGXcipNXq+3aU41VoUYrKdXJbOZXs6nk8tJn8UTYAlNfIO8vhLRhSbjUll6qdhoLjsyD+paP94qKu90gudfd0Kt1vib7RTdh91XbVIze3W6sK7AJjHlbKXLYEv8fwYX1QYt45TzSce7puHfYCLXvnr72Bx7eEm3f7Txmf2HyENSes0DtWdsQGQube6TVoDJne9+sLi5P3sz//uauZzzcg9MxtvfAn5HDX2Hr0PB08FsqbU72EDYsIybTVZS9HpJMM+zXCgKkoomUSXzFZbpf23UZcX20YC4uC2+utV0zKrnScbFfuO8O15f7eknfg1vPmsE9dF/IDbVNhq/MtMqJrv4HHXFzUeG3Mc26SOlHDallpPE8etBSX799uT/n7HDOOXwEfCR8Kvgo+Gj4GPhY+LhDEB/wgXRCPIBbWXW0zx/CFyEzqgKZkU9T467X4HoNrtfgeg2u1+B6Da7X4HoNrtfgOijZQ9Am8IF0edeHDIw7TQTBwCY8JpE7q9/Ue3x9CXrh8d7l1fp6ffX7eu/JbLb4/pvJEm6/yUSc7HbcdRQdbbfD13NHZ1J7EwjbPNqam2vLimMgDVa1TfqKUVWbT1NPzwtVelwnzjusF11QL020qOo6aOKWtOfrA36ObwjdMrQX34yWHO+GFk3eZtvMDfOmwX0WPRepSsRt82iT/0c3Y+7u5WZsCm7G7ZVPF8n36JhPU48RhDuqbm/q8Ne9jN1Nwdi9ia9UqozJ/u7Aws4nyVhYr57uRyChNQJuLQT13clR2rx32zNhDaU909SMtpIWlcT+ulPCNo82E9GENtJWs/9inmaHcEQ9hJ8Pjvx/COnagqZKp0e6fJo65VDdOFTvHIKu8mHtHaoTDg0pZbN7sUQTeUltxsJNp7hpjpsWuGmJm+qkaXiZrqDlNJPYFb4hL+7KanbFxUkVfwtWvBGSrurvtmoI/tefwkxN3C/f7kkp3nXVuqUSO+1OCducsC0I25Kw1WnbcDc2y2/Bks3kr7rLHqyCMka9b0/2QxUcXQhqAiEWj+3xWjz/YXY0g6pNDZ/bzIndTrsqiGm+5iYr/Ly332UVy8TmggK7b7zTShZMvEuOdYdfdX9Kexolbz/9iSxjjSVju/GAQ0n2oNfdYw/PqyU5miHPx9iHUMS1+FDDTVBGxolxyQ8Atnx4RRZMc8s/ZLeHHEy1ftVKbkpWzXgLAEvPfaGX2atTUVCnIl2neGhWJvdX7Oo0zjhLAXUqCuq0YKeElNmW+3W3SrQnvEsbRqJnseSbtmEUgV59y4L6lun6lnh9y0kblv36jtsVpIT6LgiRyoJ5NfkHbzyQ1Z1bM1bd1bA1p+fh5ugF92q3YD5LVunaxVfPyeRa/13txrX+soLarQpqt2DVvlS51vx1lk6357lDI5ZIHDKWfNNGjF54r5pVQTWrdDXjqyOGUMTFy7tqjkvqpYJqVgXVXLBAXmaftHK3lTFS37XRSoVUr75d9eYflxJ9JmtXJ2tXoJvOazldmq77tRuXpksNtasLardgkblE5xa+6mrW9jR3acMaqWRzu0rOz6ks2jR0LSMLnS1ey2bShk2/luNCZ2mglk1BLRcsWZb0kuWvtPRUjpfi3tvSU3m71Z0ezn/bpad3uKZb5kOWnrbVeOMWhXGs16AKVoBLZAW4wxvUdAW47TeouAJcWmhQtqBBFazlli6vZO62Oac9g+o1pnvanFOx2xHI/RtvzrnDNd0yH9aY3O0aE8qvXmtyBa0JWbJt8NY05l3Yh7NrTXHJtnQHAa5ca6pYPspatZOqyPwkFv0OU4zYHDsWBo9TkvjEevTGJCbWCducsC0I25Kw1dFmY2gb6qCW/Biq6XCv6w4O20OZmuDZkE7F0ZAObjrFTXPctMBNS9xUJ03DyxQFD0ZtJr9yz9n7+eXfZiEMGVu7jM/Sa7/js/QGj9vlN3w2n1RM0wsyWneTj1NFTXPctMBNS9xUtybiMXuVzAd5qvYpI1SARodpA8mYsiezo9kAc/rJpM9i8cRD+Bi3/bVadDyICgbFc+HrxrqyksHqVM8fy6RiRDENESPKxoQKzlPHNIT2D2v5btbLDhlTFTCmmdzC4q/QSh/y2X6l2BflzIMkKdrpO9dR9/cnMLRQenS3btNRyzvmqDe9+mkfPpwKo+az1ztv49CvqtoFJPDXAVxNpqNXBZjSj9wImO4vnv/9QQbY7rEOO2ChOVk3BlYVAJt/oHGbJA1sNnu987YDVnXAKgB2+iDDIbC6AFj66b0NsD8+fZ0DVhcyNr8gaY661ANWE8Bms9c7bztgdQesBmAzq00qUwAs/TgED6yY7VvrvkBflMbUTMkqpLBqjKkpwBTzpoepITDNZq933naYmg5TA5iaDKa2AFObJauc7QOeXwDXNKY2wVMFI6YxprYAU8ybHqaWwDSbvd5522FqO0wtYGozmBY8R7iit3kHTNuelQTWJXpWZc2kA3AFwGIu9YB1BLDZ7PXO2w5Y1wELSrHKaBNVsDpIsTxZ256VArYtJsvYNh0JLOrSDtg2SRLYfPZ6520EVrEIrGIHcDUZYHkBsPRO1gDst/7/arbPNWM4tHzKWW2sG3O2TUdDiznVg5YT0Gaz12q8wxMK7KDlAC3PQCsKoBVZznpQmTY4qGLKV8W54GNQRQGomDs9UAUBajZ7vfO2A1V0oAoAVWRALVCQin7gZgB1BahWQuCoyilVK25UNUZVFqCK+dNDVRKoZrPXO287VGWHqgRUMwvEVYHKUnmVpZqelbOKp0FNySwj2aRrLZBZqDs9UAmZlc9eq4nMUp3MUiCzVEZmqQKZpfIyS7ftH0c1obGUdMaOUS3QWKg/PVQJjZXPXquJxlKdxlKgsVRGY6kCjaXyGkvHDgCHNaGwqspZMYa1QGGhDvVgJRRWPnutJgpLdQpLgcJSGYWlChSWyissztsugLsqjWpKY1nJJ6gWaCzUnx6qhMbKZ6/VRGOpTmMp0Fgqo7FUgcZSeY3FZTuwwmFNyCzDnJ6MrApkFupQD1ZCZuWz12ois1QnsxTILJWRWapAZqm8zOLyF8DVKUXgmlBZxjA1oWuBykI96uFKqKx89lpNVJbqVJYClaUyKksXqCydU1mvTwFUxdQXwXh6bKUTCssyq8eg6gKFhbqzA1UTCiufvdYThaU7haVBYemMwtIFCkvnFFYA1XJGgJrQVlZpMx5b6QJthbrTA5XQVvnstZ5oK91pKw3aSme0lS7QVjqnrTyoouKOADWhrbi2ejLNqgvEFepPD1VCXOWz13oirnQnrjSIK50RV7pAXGlaXL18s3jIDbftDBbXDnpXy3Ua3oTI0pLLscjSBSIL9auHLiGy8tlrPRFZuhNZGkSWzogsXSCydFZktcBqaYC2jKVxTeksIO2EtQU6C/Wohyuhs/LZaz3RWbrTWRp0ls7oLF2gs3RWZ7W4GqsIXBNKSzjJx0pLFygt1KMeroTSymev9URp6U5paVBaOqO0dIHS0lmlFXF1jMA1FcxK4VogtVCPergSUiufvdYTqaU7qaVBaumM1NIFUktnpVaLqxXVF+4wXFNiK4VrgdhCPerhSoitfPZaT8SW7sSWBrGlM2JLF4gtnRVbPVzxYUFCbclK+yfwD3EtUFuoRz1cCbWVz17ridrSndrSoLZ0Rm3pArWls2or4lopgq8JtSUcpB7jWqC2UI96uBJqK5+91hO1pTu1pUFt6YzaMgVqy2RjWhFXJfD+1SQEl+TKyRGupkBwoR7tcDWE4Mpnr81EcJlOcBkQXCYjuEyB4DLZkFaLq6OErEloLgnjrHH/ago0F+pRD1dCc+Wz12aiuUynuQxoLpPRXKZAc5lsPCviSo1fTUJ2ycpMcS1QXahHPVwJ1ZXPXpuJ6jKd6jKgukxGdZkC1WWyIa0ouBizaFDLJPSWgBzj1VimQG+hHvVwJfRWPnttJnrLdHrLgN4yGb1lCvSWKdVbIQQDd/g0rgm9peG+NZ57MQV6C/Wohyuht/LZazPRW6bTWwb0lsnoLVOgt0yp3iJD2yalt1J8LdBbqEc9XAm9lc9em4neMp3eMqC3TEZvmQK9ZUr1VsAV72ATgkvBgGBC2ALBhbrUA5YQXPnstZkILtMJLgOCy2QElykQXKZUcDVxGCUQYBOKi2vLxQTZAsmF+tRDlpBc+ey1mUgu00kuA5LLZCSXKZBcplRyAU7si1YIsAnJpZgw46lCUyC5UI96uBKSK5+9NhPJZTrJZUBymYzkMgWSy5RKrjDFzRXWFaTWETJT2fH6IVMgulCfesgSoiufvTYT0WU60WVAdJmM6LJxuxmx+8e2+gXZgYJtNxvuBrr9/rJ4+tSTOQnbnLAtCNuSsNXR1sDd7i+rqmOoh8H+suZQBvr8/jKL7y/DTae4aY6bFrhpiZvqpGl4mQX7y2zZ/rIfVh9WZ7PdQ/+AO0oNdpJNX+OAcArfIYab5rhpgZuWuKm22R1itkAM2fvdIRaL/yN2iMVzfc0dYrFMaodYTHOXHWIF56ljmnvcIWYLZJ6lZd4PPz/kSSYkdF3FnB3H0WyBrkNd6FUKoevy2Ws70XW203UWdJ3N6DpboOssres8kLP9F4uH6Q02NqHnqspM5iFtgZ5DPenhSei5fPbaTvSc7fScBT1nM3rOFug5S+s5wFMmgUytTTRCTIAs0G+oCz0gCf2Wz17biX6znX6zoN9sRr/ZAv1maf3mgQxvYWj6UCmNTvf9Cf1mXKXHg2FbIN9Ql3rAEvItn722E/lmO/lmQb7ZjHyzBfLN0vLth5+voclrofwoBW6DaVRT8TItJvNjtkC8of70UCXEWz57bSfizXbizYJ4sxnxZgvEm6XFW4OqyaCaUG5SKznhaoFwQ/3poUoIt3z22k6Em+2EmwXhZjPqwRVEyxwdLWtQtTSqLhUr02Yy0eAKYmWoPztUHREry2evd95GVF0XK3PsAK4mg2pBrMzRsbIG1eamn+sIXCpg5pQYB3ZcQcAMdasHLhEwy2ev3SRg5rqAmeMAbiZg5goCZo4OmA3AzfQHLhU1c3ayVtkVRM1Qt3rgElGzfPbaTaJmTnTgCgA3EzVzBULR0VGzAbi5biG1H4wxN94b7gpCZ6hbPXCJ0Fk+e+0moTMnO3AlgJsJnbkCTeVymur6ocxSNiWwxDTQ6woEFupPD1VCYOWz124isFwnsBwILJcRWK5AYLmcwPKo5riaklmS8fHAwBXILNSfHqqEzMpnr91EZrlOZjmQWS4js1yBzHIZmeUfEGOcXwguqiSiCb0lmZ4spHMFegv1pYcoobfy2Ws30Vuu01sO9JbL6C1XoLdcRm8Bog85C6D6F/0lQU2tThTajld7uQKthbrTA5XQWvnstZtoLddpLQday2W0livQWi6jtfwTTDi3HOdpQmhVTk/2grsCoYU604OUEFr57LWbCC3XCS0HQstlhJYrEFouI7QAUs/TgCpK1KTQYm4iCQqEFupPD1VCaOWz124itFwntBwILZcRWoBCHtaQiKaqnHGlGUrVWMJw0Qznk4dBxIQkrLg/O1xjmiSwBQXUPZe7Z66yTm7Bnwf+qnLoFiiukIhmrQys9QBjrI2FDIcBSopx/xoTZgDOy62YBgE4L7h6Lu8A5juAuQc4I7o4K1BdIRFF3zrcv0wa14TOqqyZbF+MCTO45pVWTIPgmtdaPZd3uIodrsLjmtFbnBW8ADwkur556BwLdN4+iN55koqiU8Y5ZVxQxiVlrDvjIJRuq2NfO4NYenssVx1VNpoe0lynw+mE7ZSwzQnbgrAtCVudto0uV5WwrxU0aLCxfWzr/MXbv892b5dplmykWdS+UCOJEmqbE7YFYVsStjraiOA5ZwVaKCS6vr/weVf+HxE/7072NQPoXaFUBL1LdJcQesmZ6i7RPQbROSvQfCERcdds2tVy+WL27HsmZvtKfQkL3VIcMdN7KOdSqfF8WkyZuYlijvXri9B/BQXUPZ93N1Gzu4kafxPNiEDOClRgSESM/nYwv/yeVX7L6heHwdy9+qAHs9SVnMJcoAhxx/owE5qwoIC65/MOZruD2XqYM8KQswJlGBJl2bx4PXvGeMPl9ALKWM7w2SuOm+k4u0Af4l71MSYUYkEBdc/lHcZuh7HzGOdkIi+RiZwVUBkwfilUIHJ4+lMC47ac4dMtTKK74CViEfWqhzGnxGK+gLrncocx34lF7sUiz4lFXiIWOS/j8UvfJwPGWJ/cljPsk5n/NwG5RDCibvVBpgRjvoC65/MO5J1g5F4w8pxg5CWCkdOvG+2BTPfIXCR6ZKZFAuQS9Yi61QeZUo/5AuqezzuQd+qRe/XIc3KFF4TrQqIyJtsAMq+QLrktaEhlwaHHmKBcEK/D/eqjTETsCgqoez7vUJY7lKVHORO247wgbhcSZan8c0vlsUqCb6ubV0FzwdQJgn81ZfmU4AVBPdzZPvREWK+ggLrn7g76agd95aHPxPY4LwjuhURZgnvhYFnz2lPJdftGkfhGkPaNIjC8TyOvpsxXYrKDPqbLgI+52wefiP4VFFD3PN6Br3bgKw9+JgTIuS6YHeD6VnNTXaXcfi4qnjk5F0UY55RxQRmXlLHujIO5KFAvx746BpNR8WCuAkx+NoobfDYKt50StjlhWxC2JWGr07bR5doSvtmi2Sjfw4owtwJM0s2rqrSuwjd0teGbCTN4wZC0TXcgXcNAbgXCPIvPX+G2OWFbELYlYaujjZq/4iW6jbvc/JW0jjfzV1WVXufbFYJPUvlZm+Z+56y72aRUW3gaWdS2IGzLaNNJZN39TxCJErknso8ruQHdexN31tIdrJiItd+fOMamt72YsEpVDWFcUMYlZayjsb3BeWfpvlSUiD5BRwhfvv3+7WwHqdCMYdANQm5CNailXlUYk+okrwnjgjIuKWM9cg9uU/vPfnz7w5vw3kF+GN47+ODb3SE5PaSmh/T0kJkestNDrj30wN8d2aFHL/GixFF1lshLQccjm3l3PXxZHBPtS+OaF2c2o5N4j/BfPDQm48LUuuI6vFCznZNuGLFLD12lVif98/hXsqcZM9GTvz/RKtXYBNXYcOOCMi4pYx2Nu8YmMrVToksFvY60qQ05qI1KjV4BGLqwpmtLo9rqR+Qdl4lAhhTG7AIZkrH0QoDTruh0PeDGBWVcUsa6MzLibi+iWnUU9q1KtFTT0GbQNCobKcwK7i83fR+jVa5qWpBt+tYqtLz2db3xfpV8aW+8mju853xK6FjLJXIaR7Nf6xM9/fuT0QqyZUzUvTl23+frvYDY/0RfIjvCwb84Fg61b46Fvw48Url+tSQiLNStNB82xr6TCoy+JFUgYZxTxgVlXFLGujN2KnDfD1aO6wLsdV7uCY3LPdx2StjmhG1B2JaErU7bRpdrSqhmSuVe29SF1rbqUUnCHRqjjsFlHG6bE7YFYVsStjraKBknWnmsUnCNkraKz3cIOLSObMVhxULFTK8bHTdTGAe0GkLCsGe2+L73IugwEpqFpgtDJuqt0btEcTq0f2S4eKlvwTqT7g7knFG9TqR3U2DNOgkWhxXcZDoZR3UyuHFOGReUcUkZ685oQifzCobSp/v7tVDf1hX7trbw4Ux+66ekvvU90oMjv56RwZ3jgKekyJA9khUTTZaoKkmrKl8zTMpmuBxu08IpM1ixcqPb+Q0WpTyLvqVk0illnFPGBWVcUsa6M9rbj2VkpnpFQecgmxG/xFcA+ed3hublTNBMooLuwX/L5rEpO38HnUSuVwlN3pzM9jn8ZY1oOvckQdJRktZz7ZI1ihvnlHFBGZeUsY5Gw3pt1bfgo33+0Nf3DRunLG+cJXErScetiHpWVbPM7A9srBXVWHHjnDIuKOOSMtad8Q6NtcpUtypprO1+MOxO/uynkmZZ1J7TzbOIIkhzbX1P3loJ45wyLijjkjLWnVEOmqsUbXOtbthcdXlzLVksKOkdYqGm/w8apqEaJm6cU8YFZVxSxroz3qFhqkzFxmAVWVvtyjxcvkxnIOPDxxS3vcrw/rF+5dDNedriEw13ShakibZXka5d3DinjAvKuKSMdWfUwyaq2iZqbthEXV52S4fLbtx2StjmhG1B2JaErU7bmss9uv64Xm9PV9vVk0cX66sP62fr8/Pr2a+b3z5BWulHJ93h2dX6faO+j4M0PUrYuD0O0cep7Zk6PlXJPFIch8FSsjx9HOLrKRuTx2EfSLJMcxw6g4StOq6r5HF2vKxYymLBYpMWztRx2C2QsAl9fCpSnvuw1HEIZKVyHdcpJPyk9XGYB05erToO96/UVUGBVbpE7o5DrDPlBVSWQGrLHoeGmLIJX6JIlgj3yOMwbkrhe1wnGQPd13GIQSTP5a8Z8RHqJO0DPw5CK+UD4GSTZwK6n6bZDpdbN1d7tGs7Tx69g9b08+r8DL7PNp+6xqT2xqbZ9T9DQUs/uPi4+Xx6tbk83Xz+9HiPNQeef7r8bftyfX29+rDuDs6vrjZX/YOr8/PN56fnq0//CD+3/7qE4+dn11vfEW6uLn47X/Enfu1lUFUn3IlDqIjuxx50hDHVo6Ohh5jHz/RX85ijHtfPH3J5GP5/7b8kC/+r8L8O/5vwv/P/Vzz8b8P/3zX/f/vix8Ub/6cKyVVIbl/d5ooX3MwWIfj/1S577e1vzrbnYA0TUjIIYM6rJsTSjQFiaATu8W2ukCGsM9jntnrQTFc101ayjYLJ5g4urJAn+5C2SSSb6IwLs6E+gtacQhrVZmqGE7KN8kjbFtoOh+CU4MLl1ebictt5nt6vND5NzOZzSOisgsVIp08eNwObdtbNtvElIYe/u7jR2G5PUAYNQDocoHEr1odI23031Fer7ZoPoh1VMz1q3S7qcYhj5td8hEGtlsac3OY6TyWfnUoJHwUfDR8DHwufr8d/dkf+L57/gLE+zCrfE9cnk9XMMktwHdx8fPOwZzMFfujnxx9Po5/pCXOUT+DCoV9JQRJhdADuYpdQay9XVx/O4A52vn4P18L+4gfeV80wuPmx3VyGruyXzRaGyOHPj+vVu/WVTwD295vNNv44asp8vd7+djl7f7Z9s/nb2bvtR5/J30E/b67+EQajT/4XUEsDBBQAAAAIABFNaFxR283yMSQAAAXtAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDIueG1s3V1rcxs3lv0rXO1URlbkER6Nl2S7KrZIdiZ2nPIjM1Vb+4GJaVs1kqmRmNizv34v0I1mP3AvIMnKTm25aIp9AfTtgwM0Di7Q/ejz5uof1x/X6+3sy8X5p+vHex+328vjo6PrXz+uL1bXf9lcrj+B5f3m6mK1hZ9XH46uL6/Wq3ch08X5kWBMH12szj7tPXkUjv109eTR5rft+dmn9U9Xs+vfLi5WV/96uj7ffH68x/figVdnHz5u/YGjJ48uVx/Wr9fbt5eQ/v3Z9s3mJzjQ2o66Mt+dXaw/XZ9tPs2u1u8f7z3lxzWXzqcJSX4+W3++7v09u/64+by8Onv3HByB62J7M3+tv2w2//Dm79/5Q/7cn9azL68vz8+CN7N/tX/Kvdl2c/l8/X77bH1+Dqer9marX7dnv69/ghyP937ZbLebi3AVcE3b1RaOvb/a/M/6U/Bofb6GxODrZUgNRTVJE7amJH8m3Nqep3Ggceg7j+U/AxTf9ZDy19b/O0KyCBUIAP+yul4/25z/7ezd9uPjPbs3e7d+v/rtfPtq87let5WifHm/bs6vw/+zz01agOTX367BmzYvOHBx9qn5Xn1p66ufXiAZRJtBjDLwCskg2wxynAFzqWozVKVnUG0GVXoG3WbQpRlMm8GMMygkg20z2DGsSHrXpg9N4qipvVD1p6vt6smjq83n2VVI7atYslhKV+nQGH71KZ56YoWEcPTsk2/Hr7dXYD2DArdP3nz3/OU3/8mFEurEf1e6ar5d+1u1v408mTUJeHPAyvbbwrfkkrnwraULx1X4FkxaX5DQ2vqCpJJSd+lms/+aZhXMSuO/Oa/8t1YCTuG9lNwFr/wRd/Lfj462gIy/jKNf4QOIdLCIBhaucVhEgEUgsPgrkMx7XDFdBc+ldA0UkoUrUk6HKzdSmITbwkinPWTgLXg925UojGEsVSKvKuG/HZMNxoq34FvT/BYN5lK32IdKk1IoG/KLWI4O5VSV4R2A8FsYJ0NdSKWbdNZ0cCYclRW3wVHBKok6On/7qkEhVKJ1UFb/qgVvajwyAQpxLUM466jVfLfHq3icD3+Hq/D5eUOXKtDFGSZOCC7IhgsEFWSgggxU8Pc9X0SwPEMtp6hljloWqGWJWuqUZXB5VXN5gugBqlBGhVLdaSmaeldMngx6AyPaXiD2Dv1GD+ZQdUIoY0ZV0ODXnFkl8EMtc9SyQC1L1FI3Fo3jp/L4qVCGIbqKG+LGFU/j1ZzJJvBSieto8MK82+FW4P+499eu9b+9DlOl/F32/Z2WzLkIF3vCnUjkrlNXO6gbna8bnbm20k41dpJOa9bvZbQL9y1IVekmW5tdtve/2JvF+2LTW6VuYi5Z5Rqt8sbi0Evr+nXm7wfHs+cvHwpOnGqOYbVjSg5NqeCCerd7oRl06ilmNCVxlmiSqSseVLvJ9tgG7bFRyylqmaOWBWpZopY6ZRlcns2z2t6gx26HY6qlX2S5GP6u7JCmw+GbgDFO882Y61pDuLOLSkbGO0Yy2aKdPWqZo5YFalmiltrmOnvXjgstDr1riIuNl8dYpYBoi8DHlsMx0QxwThRz2hYji/u24ZA01QO0ReLEmgztEqUs2lIUXsp4LJxzbNkWqYkiZWC4cM6JkiLrtkji1sCsYc1wlpmuNoiRpO/NAlkcobbYDW+2MCoXu1H6882Hv63Oz2Mz+68L8d8phrVnAR4/efT+yfeL/e9f//j2xdP5q/3X8+9ePav3gTbV3uFSPXhw+OfUGJwae//5P57pW+VSDx4dvX/y6Hfw9Pc+kaO3rvV2/urVy1f7Pz9/+fKHtz/ttyczPPRMwsr2JOn79JCgf/6PPz09/tPbQ3O4+O756/mDw729tA/z1gfBBj4AkAfg3iFL51rEXDzRDUWbiHUAqb85355882F7srd3CL8O42maU6TPUcdyZCwHSn7M4EoO4Y8D7+Ft6kJOzjakMi+gMs+NrSrBRHP3AK3HF5HEi/PN5urpZnX1bhYPElTmfSrf5lINwjp+36yzedbxIev2TyUHWn3/497hjoIcKMiBHzgLOcFCPmEhH7CQ91jICRbyEQv7ngIleaQkP/Ae346SKaCGtBQFtBQ3oqVI0VLkaSnui5bivmkJ1ZzlpSjgpQBeCpKXguClmPBSDHgperwUBC8FzUsReSkOvMf3xUtZwEuZ46Ww4c4vGOcm0W0+W59Bpg8+IHFe0nHKuzLUIgyV/wYdp5wQVE4IKoGgkiSoJAgqJwSVA4LKHkElQVA5IajsE1RGgsoD7/F9EbQqIGh1K4KKNEELutDqvgha3TdBXZ6gVQFBKyBoRRK0IghaTQhaDQha9QhaEQStaIJWkaDVgff4vgiqCgiam6gEd1yQ3uBP0E7aqN3k3fPVxdmn1Xb97ul6dbEjp0yTU92VnA4hp7pvcqo8OVVa7Cjgo8LJqAgyqgkZ1YCMqkdGRZBRTcSOivxTB97D+xA7uoB7ualPHxJt6qZysldHnW4P3WK2T9R3pR1nCO/0vY8rWZ54etIrqkmvqIGFmuwVNUFEPSGiHhBR94ioCSLqSa+o+r2ijqzUB97j++oVTQEzTfa2XUW9w2zHyFebzftXq/fb9VWWk2Y0mbTgpgdpc+DxXjM+kCf7QqoHe7eaIIKB/q2y8QddpYI3X5XTVcvpx4LbP/qabpkPm2VrqlFUgw6gGRogrdVMWqubtFYDrdWQrbUtJTVVH21i4NMS96mOOWTKJ8hpYrs0B963+2qXtqBdWrJdegHnV70EbkLD0V3L5ItZmIr46+bsepttnPbONwyJ8MWm+cJxvtgJX8yELxb4Ykm+tKVM5/p7vLFp3qC+1TGHTPkGOW3kjT3wPt4Xb1wBbxzJG4HyRtyIN27Sqbtxp+6+Tqd+u2ABV/fXqas7d+q3v6Zb5sM6dZdupAJvpG7SSO2kkTpopI5spI7o1F26caI+1TGHTPkEOV1snO7A+3ZPjVMUhO9ELnwXw3XCN7+uafqY3eJqdbHONUvB7tydI/PKbckTpmiUKYJNmKLHTBHsAEqmmBJLobpzwdKMQX2rxTgo1vdtHwpsGQP+eR/vizEFUTKRi5L5Pku3fRbvOdQw58366mr1a543d46ScWS+Qtx7mIzLrG4UozhZZB8H9qH3/IUg4mJiEhcTg7iY6MXFBBEXE+O4GJQcuccPvIf3MGEhCsJggg6DQVNp6icu5GMnvWU9L2enMJDoSBe0IxfjJdUt9e4cCRPIlIW4/1CYzlNPpKkngHroHW0hiNCXmIS+xCD0JXqhL0GEvsQ49AUlR+qJA+/hfVBP5leniTbkQfV50jZ9nWwmbBm34uTVarv2VGNVqNFKSnUym/nVbCq5vPRZPBG2wNQXyPsLIW1YEi61pZeqncaCI/OgvuXjvaLibjdI7nU39Gqdr8l+0U3YfdU2FaN3txvrCmwCY95Wihy2xP9HcGF90CJeOY90nHs67h02Qu27p6/9gYe3RNt3O4/ZX5g8BLXnLFB71jZExsLmHmk1qMzZ3jeri8uTN/O/v7nrGQ/34HSM7T3wZ+TwV9g6NDwd/JZKm5M9hA3LiMl0FWWvhyTTDPu1ggCpaCJlEl9xme7Xdl1GXB8tmIvLwptrbdeMSq50XOwX7rvD9eW+XtL34NazZnAP3RdyQ22T4SszrXKiq/9BR9xcVPhtTLMuUvpRQ2oZaTyPHrTU129f7M85O5xzDh8BHwmfCj4KPho+Bj4WPu4QxAd8IJ0QD+BWVh3t84fwRciMqkBm5NPUuOs1uF6D6zW4XoPrNbheg+s1uF6D6zW4Dkr2ELQJfCBd3vUhA+NOE0EwsAmPSeTO6jf1Hl9fgl54vHd5tb5eX/2+3nsymy1++GayhNtvMhEnux13HUVH2+3w9dzRmdTeBMI2j7bm5tqy4hhIg1Vtk75iVNXm09TT80KVHteJ8w7rRRfUSxMtqroOmrgl7fn6gJ/jG0K3DO35N6Mlx7uhRZO32TZzw7xpcJ9Fz0WqEnHbPNrk/9HNmLt7uRmbgptxe+XTRfI9OubT1GME4Y6q25s6/HUvY3dTMHZv4iuVKmOyvzuwsPNJMhbWq6f7EUhojYBbC0F9d3KUNu/d9kxYQ2nPNDWjraRFJbG/7pSwzaPNRDShjbTV7L+Yp9khHFEP4eeDI/8fQrq2oKnS6ZEun6ZOOVQ3DtU7h6CrfFh7h+qEQ0NK2exeLNFEXlKbsXDTKW6a46YFblripjppGl6mK2g5zSR2hW/Ii7uyml1xcVLF34IVb4Skq/q7rRqC//WnMFMT98u3e1KKd121bqnETrtTwjYnbAvCtiRsddo23I3N8luwZDP5q+6yB6ugjFHv25P9UAVHF4KaQIjFY3u8Ft//ODuaQdWmhs9t5sRup10VxDRfc5MVft7b77KKZWJzQYHdN95pJQsm3iXHusOvuj+lPY2St5/+RJaxxpKx3XjAoSR70OvusYfn1ZIczZDnY+xDKOJafKjhJigj48S45AcAWz68IgumueUfsttDDqZav2olNyWrZrwFgKXnvtDL7NWpKKhTka5TPDQrk/srdnUaZ5ylgDoVBXVasFNCymzL/bpbJdoT3qUNI9GzWPJN2zCKQK++ZUF9y3R9S7y+5aQNy359x+0KUkJ9F4RIZcG8mvyDNx7I6s6tGavuatia0/Nwc/SCe7VbMJ8lq3Tt4qvnZHKt/65241p/WUHtVgW1W7BqX6pca/46S6fb89yhEUskDhlLvmkjRi+8V82qoJpVuprx1RFDKOLi5V01xyX1UkE1q4JqLlggL7NPWrnbyhip79popUKqV9+uevOPS4k+k7Wrk7Ur0E3ntZwuTdf92o1L06WG2tUFtVuwyFyicwtfdTVre5q7tGGNVLK5XSXn51QWbRq6lpGFzhavZTNpw6Zfy3GhszRQy6aglguWLEt6yfJXWnoqx0tx723pqbzd6k4P57/t0tM7XNMt8yFLT9tqvHGLwjjWa1AFK8AlsgLc4Q1qugLc9htUXAEuLTQoW9CgCtZyS5dXMnfbnNOeQfUa0z1tzqnY7Qjk/o0359zhmm6ZD2tM7naNCeVXrzW5gtaELNk2eGsa8y7sw9m1prhkW7qDAFeuNVUsH2Wt2klVZH4Si36HKUZsjh0Lg8cpSXxiPXpjEhPrhG1O2BaEbUnY6mizMbQNdVBLfgzVdLjXdQeH7aFMTfBsSKfiaEgHN53ipjluWuCmJW6qk6bhZYqCB6M2k1+55+z9/OJvsxCGjK1dxmfptd/xWXqDx+3yGz6bTyqm6QUZrbvJx6mipjluWuCmJW6qWxPxmL1K5oM8VfuUESpAo8O0gWRM2ZPZ0WyAOf1k0mexeOIhfIzb/lotOh5EBYPiufB1Y11ZyWB1quePZVIxopiGiBFlY0IF56ljGkL7h7V8N+tlh4ypChjTTG5h8VdopQ/5bL9S7Ity5kGSFO30neuo+/sTGFooPbpbt+mo5R1z1Jte/bQPH06FUfPZ6523cehXVe0CEvjrAK4m09GrAkzpR24ETPcX3//9QQbY7rEOO2ChOVk3BlYVAJt/oHGbJA1sNnu987YDVnXAKgB2+iDDIbC6AFj66b0NsC+fvs4BqwsZm1+QNEdd6gGrCWCz2eudtx2wugNWA7CZ1SaVKQCWfhyCB1bM9q11X6AvSmNqpmQVUlg1xtQUYIp508PUEJhms9c7bztMTYepAUxNBlNbgKnNklXO9gHPL4BrGlOb4KmCEdMYU1uAKeZND1NLYJrNXu+87TC1HaYWMLUZTAueI1zR27wDpm3PSgLrEj2rsmbSAbgCYDGXesA6Aths9nrnbQes64AFpVhltIkqWB2kWJ6sbc9KAdsWk2Vsm44EFnVpB2ybJAlsPnu98zYCq1gEVrEDuJoMsLwAWHonawD2W/9/NdvnmjEcWj7lrDbWjTnbpqOhxZzqQcsJaLPZazXe4QkFdtBygJZnoBUF0IosZz2oTBscVDHlq+Jc8DGoogBUzJ0eqIIANZu93nnbgSo6UAWAKjKgFihIRT9wM4C6AlQrIXBU5ZSqFTeqGqMqC1DF/OmhKglUs9nrnbcdqrJDVQKqmQXiqkBlqbzKUk3PylnF06CmZJaRbNK1Fsgs1J0eqITMymev1URmqU5mKZBZKiOzVIHMUnmZpdv2j6Oa0FhKOmPHqBZoLNSfHqqExspnr9VEY6lOYynQWCqjsVSBxlJ5jaVjB4DDmlBYVeWsGMNaoLBQh3qwEgorn71WE4WlOoWlQGGpjMJSBQpL5RUW520XwF2VRjWlsazkE1QLNBbqTw9VQmPls9dqorFUp7EUaCyV0ViqQGOpvMbish1Y4bAmZJZhTk9GVgUyC3WoByshs/LZazWRWaqTWQpklsrILFUgs1ReZnH5C+DqlCJwTagsY5ia0LVAZaEe9XAlVFY+e60mKkt1KkuBylIZlaULVJbOqazXpwCqYuqLYDw9ttIJhWWZ1WNQdYHCQt3ZgaoJhZXPXuuJwtKdwtKgsHRGYekChaVzCiuAajkjQE1oK6u0GY+tdIG2Qt3pgUpoq3z2Wk+0le60lQZtpTPaShdoK53TVh5UUXFHgJrQVlxbPZlm1QXiCvWnhyohrvLZaz0RV7oTVxrElc6IK10grjQtrl68WTzkhtt2BotrB72r5ToNb0JkacnlWGTpApGF+tVDlxBZ+ey1nogs3YksDSJLZ0SWLhBZOiuyWmC1NEBbxtK4pnQWkHbC2gKdhXrUw5XQWfnstZ7oLN3pLA06S2d0li7QWTqrs1pcjVUErgmlJZzkY6WlC5QW6lEPV0Jp5bPXeqK0dKe0NCgtnVFaukBp6azSirg6RuCaCmalcC2QWqhHPVwJqZXPXuuJ1NKd1NIgtXRGaukCqaWzUqvF1YrqC3cYrimxlcK1QGyhHvVwJcRWPnutJ2JLd2JLg9jSGbGlC8SWzoqtHq74sCChtmSl/RP4h7gWqC3Uox6uhNrKZ6/1RG3pTm1pUFs6o7Z0gdrSWbUVca0UwdeE2hIOUo9xLVBbqEc9XAm1lc9e64na0p3a0qC2dEZtmQK1ZbIxrYirEnj/ahKCS3Ll5AhXUyC4UI92uBpCcOWz12YiuEwnuAwILpMRXKZAcJlsSKvF1VFC1iQ0l4Rx1rh/NQWaC/WohyuhufLZazPRXKbTXAY0l8loLlOguUw2nhVxpcavJiG7ZGWmuBaoLtSjHq6E6spnr81EdZlOdRlQXSajukyB6jLZkFYUXIxZNKhlEnpLQI7xaixToLdQj3q4Enorn702E71lOr1lQG+ZjN4yBXrLlOqtEIKBO3wa14Te0nDfGs+9mAK9hXrUw5XQW/nstZnoLdPpLQN6y2T0linQW6ZUb5GhbZPSWym+Fugt1KMeroTeymevzURvmU5vGdBbJqO3TIHeMqV6K+CKd7AJwaVgQDAhbIHgQl3qAUsIrnz22kwEl+kElwHBZTKCyxQILlMquJo4jBIIsAnFxbXlYoJsgeRCfeohS0iufPbaTCSX6SSXAcllMpLLFEguUyq5ACf2RSsE2ITkUkyY8VShKZBcqEc9XAnJlc9em4nkMp3kMiC5TEZymQLJZUolV5ji5grrClLrCJmp7Hj9kCkQXahPPWQJ0ZXPXpuJ6DKd6DIgukxGdNm43YzY/WNb/YLsQMG2mw13A91+f1k8ferJnIRtTtgWhG1J2Opoa+Bu95dV1THUw2B/WXMoA31+f5nF95fhplPcNMdNC9y0xE110jS8zIL9ZbZsf9mPqw+rs9nuoX/AHaUGO8mmr3FAOIXvEMNNc9y0wE1L3FTb7A4xWyCG7P3uEIvF/xE7xOK5vuYOsVgmtUMsprnLDrGC89QxzT3uELMFMs/SMu/Hnx/yJBMSuq5izo7jaLZA16Eu9CqF0HX57LWd6Drb6ToLus5mdJ0t0HWW1nUeyNn+88XD9AYbm9BzVWUm85C2QM+hnvTwJPRcPnttJ3rOdnrOgp6zGT1nC/ScpfUc4CmTQKbWJhohJkAW6DfUhR6QhH7LZ6/tRL/ZTr9Z0G82o99sgX6ztH7zQIa3MDR9qJRGp/v+hH4zrtLjwbAtkG+oSz1gCfmWz17biXyznXyzIN9sRr7ZAvlmafn248/X0OS1UH6UArfBNKqpeJkWk/kxWyDeUH96qBLiLZ+9thPxZjvxZkG82Yx4swXizdLirUHVZFBNKDeplZxwtUC4of70UCWEWz57bSfCzXbCzYJwsxn14AqiZY6OljWoWhpVl4qVaTOZaHAFsTLUnx2qjoiV5bPXO28jqq6LlTl2AFeTQbUgVuboWFmDanPTz3UELhUwc0qMAzuuIGCGutUDlwiY5bPXbhIwc13AzHEANxMwcwUBM0cHzAbgZvoDl4qaOTtZq+wKomaoWz1wiahZPnvtJlEzJzpwBYCbiZq5AqHo6KjZANxct5DaD8aYG+8NdwWhM9StHrhE6CyfvXaT0JmTHbgSwM2EzlyBpnI5TXX9UGYpmxJYYhrodQUCC/WnhyohsPLZazcRWK4TWA4ElssILFcgsFxOYHlUc1xNySzJ+Hhg4ApkFupPD1VCZuWz124is1wnsxzILJeRWa5AZrmMzPIPiDHOLwQXVRLRhN6STE8W0rkCvYX60kOU0Fv57LWb6C3X6S0Hestl9JYr0Fsuo7cA0YecBVD9i/6SoKZWJwptx6u9XIHWQt3pgUporXz22k20luu0lgOt5TJayxVoLZfRWv4JJpxbjvM0IbQqpyd7wV2B0EKd6UFKCK189tpNhJbrhJYDoeUyQssVCC2XEVoAqedpQBUlalJoMTeRBAVCC/WnhyohtPLZazcRWq4TWg6ElssILUAhD2tIRFNVzrjSDKVqLGG4aIbzycMgYkISVtyfHa4xTRLYggLqnsvdM1dZJ7fgzwN/VTl0CxRXSESzVgbWeoAx1sZChsMAJcW4f40JMwDn5VZMgwCcF1w9l3cA8x3A3AOcEV2cFaiukIiibx3uXyaNa0JnVdZMti/GhBlc80orpkFwzWutnss7XMUOV+FxzegtzgpeAB4SXd88dI4FOm8fRO88SUXRKeOcMi4o45Iy1p1xEEq31bGvnUEsvT2Wq44qG00Paa7T4XTCdkrY5oRtQdiWhK1O20aXq0rY1woaNNjYPrZ1/vzt32e7t8s0SzbSLGpfqJFECbXNCduCsC0JWx1tRPCcswItFBJd31/4vCv/j4ifdyf7mgH0rlAqgt4luksIveRMdZfoHoPonBVovpCIuGs27Wq5fD579gMTs32lvoSFbimOmOk9lHOp1Hg+LabM3EQxx/r1Rei/ggLqns+7m6jZ3USNv4lmRCBnBSowJCJGfzuYX/zAKr9l9YvDYO5efdCDWepKTmEuUIS4Y32YCU1YUEDd83kHs93BbD3MGWHIWYEyDImybF68nj1jvOFyegFlLGf47BXHzXScXaAPca/6GBMKsaCAuufyDmO3w9h5jHMykZfIRM4KqAwYvxAqEDk8/SmBcVvO8OkWJtFd8BKxiHrVw5hTYjFfQN1zucOY78Qi92KR58QiLxGLnJfx+IXvkwFjrE9uyxn2ycz/m4BcIhhRt/ogU4IxX0Dd83kH8k4wci8YeU4w8hLByOnXjfZApntkLhI9MtMiAXKJekTd6oNMqcd8AXXP5x3IO/XIvXrkObnCC8J1IVEZk20AmVdIl9wWNKSy4NBjTFAuiNfhfvVRJiJ2BQXUPZ93KMsdytKjnAnbcV4QtwuJslT+uaXyWCXBt9XNq6C5YOoEwb+asnxK8IKgHu5sH3oirFdQQN1zdwd9tYO+8tBnYnucFwT3QqIswb1wsKx57ankun2jSHwjSPtGERjep5FXU+YrMdlBH9NlwMfc7YNPRP8KCqh7Hu/AVzvwlQc/EwLkXBfMDnB9q7mprlJuPxcVz5yciyKMc8q4oIxLylh3xsFcFKiXY18dg8moeDBXASY/G8UNPhuF204J25ywLQjbkrDVadvocm0J32zRbJTvYUWYWwEm6eZVVVpX4Ru62vDNhBm8YEjapjuQrmEgtwJhnsXnr3DbnLAtCNuSsNXRRs1f8RLdxl1u/kpax5v5q6pKr/PtCsEnqfysTXO/c9bdbFKqLTyNLGpbELZltOkksu7+J4hEidwT2ceV3IDuvYk7a+kOVkzE2u9PHGPT215MWKWqhjAuKOOSMtbR2N7gvLN0XypKRJ+gI4Qv3v7wdraDVGjGMOgGITehGtRSryqMSXWS14RxQRmXlLEeuQe3qf1nL9/++Ca8d5AfhvcOPvh2d0hOD6npIT09ZKaH7PSQaw898HdHdujRS7wocVSdJfJS0PHIZt5dD18Wx0T70rjmxZnN6CTeI/wXD43JuDC1rrgOL9Rs56QbRuzSQ1ep1Un/PP6V7GnGTPTk70+0SjU2QTU23LigjEvKWEfjrrGJTO2U6FJBryNtakMOaqNSo1cAhi6s6drSqLb6EXnHZSKQIYUxu0CGZCy9EOC0KzpdD7hxQRmXlLHujIy424uoVh2FfasSLdU0tBk0jcpGCrOC+8tN38dolauaFmSbvrUKLa99XW+8XyVf2huv5g7vOZ8SOtZyiZzG0ezX+kRP//5ktIJsGRN1b47d9/l6LyD2P9GXyI5w8C+OhUPtm2PhrwOPVK5fLYkIC3UrzYeNse+kAqMvSRVIGOeUcUEZl5Sx7oydCtz3g5XjugB7nZd7QuNyD7edErY5YVsQtiVhq9O20eWaEqqZUrnXNnWhta16VJJwh8aoY3AZh9vmhG1B2JaErY42SsaJVh6rFFyjpK3i8x0CDq0jW3FYsVAx0+tGx80UxgGthpAw7Jktfui9CDqMhGah6cKQiXpr9C5RnA7tHxkuXupbsM6kuwM5Z1SvE+ndFFizToLFYQU3mU7GUZ0MbpxTxgVlXFLGujOa0Mm8gqH06f5+LdS3dcW+rS18OJPf+impb32P9ODIr2dkcOc44CkpMmSPZMVEkyWqStKqytcMk7IZLofbtHDKDFas3Oh2foNFKc+ibymZdEoZ55RxQRmXlLHujPb2YxmZqV5R0DnIZsQv8RVA/vmdoXk5EzSTqKB78N+yeWzKzt9BJ5HrVUKTNyezfQ5/WSOazj1JkHSUpPVcu2SN4sY5ZVxQxiVlrKPRsF5b9S34aJ8/9PV9w8YpyxtnSdxK0nErop5V1Swz+wMba0U1Vtw4p4wLyrikjHVnvENjrTLVrUoaa7sfDLuTP/uppFkWted08yyiCNJcW9+Tt1bCOKeMC8q4pIx1Z5SD5ipF21yrGzZXXd5cSxYLSnqHWKjp/4OGaaiGiRvnlHFBGZeUse6Md2iYKlOxMVhF1la7Mg+XL9MZyPjwMcVtrzK8f6xfOXRznrb4RMOdkgVpou1VpGsXN84p44IyLilj3Rn1sImqtomaGzZRl5fd0uGyG7edErY5YVsQtiVhq9O25nKPrj+u19vT1Xb15NHF+urD+tn6/Px69uvmt0+QVvrRSXd4drV+36jv4yBNjxI2bo9D9HFqe6aOT1UyjxTHYbCULE8fh/h6ysbkcdgHkizTHIfOIGGrjusqeZwdLyuWsliw2KSFM3UcdgskbEIfn4qU5z4sdRwCWalcx3UKCT9pfRzmgZNXq47D/St1VVBglS6Ru+MQ60x5AZUlkNqyx6EhpmzClyiSJcI98jiMm1L4HtdJxkD3dRxiEMlz+WtGfIQ6SfvAj4PQSvkAONnkmYDup2m2w+XWzdUe7drOk0fvoDX9vDo/g++zzaeuMam9sWl2/c9Q0NIPLj5uPp9ebS5PN58/Pd5jzYHvP13+tn2xvr5efVh3B+dXV5ur/sHV+fnm89Pz1ad/hJ/bf13C8fOz663vCDdXF7+dr/gTv/YyqKoT7sQhVET3Yw86wpjq0dHQQ8zjZ/qrecxRj5+/fCj4IfwvZfhfhf+t/796fhuvF9zMFiGA/9VcX3v7m7PtOVjDpJIMIpbzqgmTdPfxGN6A+3SbK2QIawX2ua0eNFNOzdSTbCNZsrkLCyvkyT6kbRLJJsLiwoymj4I1p5BGtZmaIYFsIzXStoW2Qxo4JbhwebW5uNx2nqf3HI1PE7P5HBI6nGAx0umTx83gpJ05s22MSMjh7y72M7bbE5QFA5AOB2jcirkhWnbfje3Varvmg4hF1UxxWreLXBzimPl1G2FgqqUxJ7e5zlPJZ6dSwkfBR8PHwMfC5+vxn92R/4vvf8RYH2aG74nrkwlnZpkluA5uPr556LKZxj70c9yPpxHM9KQ3yidw4dCvhiCJMDoAd6JLqLUXq6sPZ3AXOl+/h2thf/GD56tmKNv82G4uQ1f2y2YLw9zw58f16t36yicA+/vNZht/HDVlvl5vf7ucvT/bvtn87ezd9qPP5O+CnzdX/wgDyif/C1BLAwQUAAAACAARTWhcjV3qdUUkAAAs7QAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQzLnhtbN19bXPcNpb1X+nVs5WRFXmEF+JNsl0VS91Nb+wkZceZqdraD524batGcmulTux5fv1egASbL7gXkGRlp7Zc7VbzAuDlwQGIgwuQTz5vrv9x83G93s6+XF58unm693G7vTo+Orr57eP6cnXz183V+hNY3m+uL1db+Hn94ejm6nq9ehcyXV4cCcb00eXq/NPesyfh2E/Xz55sft9enH9a/3Q9u/n98nJ1/c/n64vN56d7fC8eeH3+4ePWHzh69uRq9WH9Zr19ewXp359vf978BAda21FX5rvzy/Wnm/PNp9n1+v3Tvef8uObS+TQhyS/n6883vb9nNx83n5fX5+9egiNwXWxv5q/1183mH9784p0/5M/9aT378ubq4jx4M/tn+6fcm203Vy/X77en64sLOF21N1v9tj3/Y/0T5Hi69+tmu91chquAa9qutnDs/fXm/68/BY/WF2tIDL5ehdRQVJM0YWtK8mfCre15Ggcah77zWP53gOK7HlL+2vp/R0gWoQIB4F9XN+vTzcXfzt9tPz7ds3uzd+v3q98vtq83n+t1WynKl/fb5uIm/D/73KQFSH77/Qa8afOCA5fnn5rv1Ze2vvrpBZJBtBnEKAOvkAyyzSDHGTCXqjZDVXoG1WZQpWfQbQZdmsG0Gcw4g0Iy2DaDHcOKpHdt+tAkjpraC1V/ttqunj253nyeXYfUvooli6V0lQ6N4Tef4rknVkgIR88/+Xb8ZnsN1nMocPvs5+9e/vjN/+NCCXXivytdNd+u/a3a30aezJoEvDlgZftt4VtyyVz41tKF4yp8CyatL0hobX1BUkmpu3Sz2X9OswpmpfHfnFf+WysBp3gx9ylc8MofcSf/9eRoC8j4yzj6DT6ASAeLaGDhGodFBFgEAou/Asm8xxXTVfBcStdAIVm4IuV0uHIjhUm4LYx02kMG3oLXs12JwhjGUiXyqhL+2zHZYKx4C741zW/RYC51i32oNCmFsiG/iOXoUE5VGd4BCL+FcTLUhVS6SWdNB2fCUVlxGxwVrJKoo/O3rxsUQiVaB2X1r1rwpsYjE6AQ1zKEs45azXd7vIrH+fB3uAqfnzd0qQJdnGHihOCCbLhAUEEGKshABX/f80UEyylqOUMtc9SyQC1L1FKnLIPLq5rLE0QPUIUyKpTqTkvR1Lti8mTQGxjR9gKxd+g3ejCHqhNCGTOqgga/5swqgR9qmaOWBWpZopa6sWgcP5XHT4UyDNFV3BI3rngar+ZMNoGXSlxHgxfm3Q63Av/Hvb92rf/tdZgq5e+y7++0ZM5FuNgT7kQid5262kHd6Hzd6My1lXaqsZN0WrN+L6NduG9Bqko32drssr3/xd4s3heb3ip1E3PJKtdolTcWh15a168zfz84nr2YP+bqlDjXHANrR5UcnFLBFfXu90Iz6NVT1GhK4izRJlOXPKh3k+2yDdplo5Yz1DJHLQvUskQtdcoyuDybp7W9RZfdjsdUy79IczH8XdkhT4fjNwGDnOabMdc1h3BrF5WMlHeMpLJFe3vUMkctC9SyRC21zfX2rh0YWhx61xAXGzCPsUoB0RaBDy6Hg6IZ4Jwo5qwtRhZ3bsMxaaoHaIvEiTUZ2yVKWbSlKLyU8WA459iyLVITRcrAcOGcEyVF1m2RxL2BWcOa8SwzXW0QQ0nfmwWyOEJusVvebWFYLnbD9JebD39bXVzEZvafl+K/UgxrzwI8fvbk/bMXi/0Xb354++r5/PX+m/l3r0/rfaBNtXe4VI8eHf4lNQinBt9/+bdTfadc6tGTo/fPnvwBnv7RJ3L01rXezl+//vH1/i8vf/zx+7c/7bcnMzz0TMLK9iTpG/WQoH/5t39/fvzvbw/N4eK7l2/mjw739tI+zFsfBBv4AEAegHuHLJ1rEXPxRDcUbSLWAaT+5mJ78s2H7cne3iH8OoynaU6RPkcdy5GxHCj5KYMrOYQ/DryHd6kLOTnbkMq8gMo8N7iqBBPN3QPEHl9EEi8uNpvr55vV9btZPEhQmfepfJdLNQjr+EOzzuZZx4es2z+THGj14oe9wx0FOVCQAz9wFnKChXzCQj5gIe+xkBMs5CMW9j0FSvJISX7gPb4bJVNADWkpCmgpbkVLkaKlyNNSPBQtxUPTEqo5y0tRwEsBvBQkLwXBSzHhpRjwUvR4KQheCpqXIvJSHHiPH4qXsoCXMsdLYcOdXzDOTaLbPF2fQ6YPPiJxUdJxyvsy1CIMlf8CHaecEFROCCqBoJIkqCQIKicElQOCyh5BJUFQOSGo7BNURoLKA+/xQxG0KiBodSeCijRBC7rQ6qEIWj00QV2eoFUBQSsgaEUStCIIWk0IWg0IWvUIWhEErWiCVpGg1YH3+KEIqgoImpupBHdckN7gT9BO2qjd7N3L1eX5p9V2/e75enW5I6dMk1Pdl5wOIad6aHKqPDlVWuwo4KPCyagIMqoJGdWAjKpHRkWQUU3Ejor8Uwfew4cQO7qAe7mpTx8TbeqmcrJXR51uD91itk/U96UdZwjv9IOPK1meeHrSK6pJr6iBhZrsFTVBRD0hoh4QUfeIqAki6kmvqPq9oo6s1Afe44fqFU0BM032tl1FvcNsx8jXm83716v32/V1lpNmNJm04KYHaXPg6V4zPpAn+0KqR3t3miCCgf6dsvFHXaWCN1+V01XL6aeC2z/7mu6YD5tla6pRVIMOoBkaIK3VTFqrm7RWA63VkK21LSU1VR9tYuDTEvepjjlkyifIaWK7NAfet4dql7agXVqyXXoB55e9BG5Cw9Fdy+SLWZiK+I/N+c022zjtvW8YEuGLTfOF43yxE76YCV8s8MWSfGlLmc7193hj07xBfatjDpnyDXLayBt74H18KN64At44kjcC5Y24FW/cpFN3407dfZ1O/W7BAq4erlNX9+7U735Nd8yHdeou3UgF3kjdpJHaSSN10Egd2Ugd0am7dONEfapjDpnyCXK62DjdgfftgRqnKAjfiVz4LobrhG9+XdP0MbvF9epynWuWgt27O0cmRdqSJ0zRKFMEmzBFj5ki2AGUTDEllkJ154KlGYP6VotxUKzv2z4U2DIG/PM+PhRjCqJkIhcl832Wbvss3nOoYc7P6+vr1W953tw7SsaR+Qrx4GEyLrO6UYziZJF9HNiH3vMXgoiLiUlcTAziYqIXFxNEXEyM42JQcuQeP/AePsCEhSgIgwk6DAZNpamfuJKPnfSW9fw4O4OBREe6oB25GK+pbql370iYQKYsxMOHwnSeeiJNPQHUQ+9oC0GEvsQk9CUGoS/RC30JIvQlxqEvKDlSTxx4Dx+CejK/Ok20IQ+qz5O26etkM2HLuBUnr1fbtacaq0KNVlKqk9nMr2ZTyfWlp/FE2ApTXyDvr4S0YU241JZeqnYWC47Mg/qWT/eKirvbILnX3dCrdb4m+0U3YfdV21SM3t1trCuwCYx5Wyly2BL/D8GF9UGLeOU80nHu6bh32Ai1756/8Qce3xFt3+08ZX9l8hDUnrNA7VnbEBkLu3uk1aAyZ3vfrC6vTn6e//3n+57xcA9Ox9jeI39GDn+FvUPD08FvqbQ52UPYsIyYTFdR9npIMs2wXysIkIomUibxFZfpfm3XZcT10YK5uC68udZ2zajkSsfFfuG+O1xg7uslfQ9uPWsG99B9ITfUNhm+MtMqJ7r6H3TEzUWF38Y06yKlHzWklpHG8+hBS33z9tX+nLPDOefwEfCR8Kngo+Cj4WPgY+HjDkF8wAfSCfEIbmXV0T5/DF+EzKgKZEY+TY27XoPrNbheg+s1uF6D6zW4XoPrNbheg+ugZA9Bm8AH0uVdHzIwbjURBAOb8JhE7qx+V+/xzRXohad7V9frm/X1H+u9Z7PZ4vtvJku4/S4TcbLbctdRdLTfDl/PHZ1JbU4gbPNoa26uLSuOgTRY1TbpK0ZVbT5NPT0vVOlxnTjvsF50Qb000aKq66CJW9Kerw/4Ob4hdMvQXn4zWnK8G1o0eZt9M7fMmwb3NHouUpWI2+bRJv+XbsbcPcjN2BTcjNsrny6S79Exn6YeIwh3VN3e1OGvBxm7m4KxexNfqVQZk/3dgYWtT5KxsF493Y9AQmsE3FoI6ruTo7R5765nwhpKe6apGW0lLSqJDXZnhG0ebSaiCW2krWb/xTzNDuGIegw/Hx35/xDStQVNlU6PdPk0dcqhunGo3jkEXeXj2jtUJxwaUspm92KJJvKS2oyFm85w0xw3LXDTEjfVSdPwMl1By2kmsSt8R17cldXsiouTKv4WrHgjJF3V323VEPw/fgozNXHDfLsnpXjXVeuWSuy0OyNsc8K2IGxLwlanbcPt2Cy/BUs2k7/qPnuwCsoY9b492Q9VcHQpqAmEWDy2x2vx4ofZ0QyqNjV8bjMndjvtqiCm+ZqbrPDz3n2XVSwTmwsK7L71TitZMPEuOdYdftX9Ke1plLz79CcSsYklY7vxgENJ9qDX3WMPz6slOZohz8fYh1DEtfhQw01QRsaJcckPALZ8eEUWTHPLP2W3hxxMtX7VSm5KVs14CwBLz32hl9mrU1FQpyJdp3hoVib3V+zqNM44SwF1KgrqtGCnhJTZlvt1t0q0J7xPG0aiZ7Hk27ZhFIFefcuC+pbp+pZ4fctJG5b9+o7bFaSE+i4IkcqCeTX5J288kNW9WzNW3dWwNafn4eboBfdqt2A+S1bp2sVXz8nkWv9d7ca1/rKC2q0Kardg1b5Uudb8dZZOt+e5RyOWSBwylnzbRoxeeK+aVUE1q3Q146sjhlDExcu7ao5L6qWCalYF1VywQF5mH7Vyv5UxUt+30UqFVK++W/XmH5cSfSZrVydrV6Cbzms5XZqu+7Ubl6ZLDbWrC2q3YJG5ROcWvupq1vY092nDGqlkc7dKzs+pLNo0dC0jC50tXstm0oZNv5bjQmdpoJZNQS0XLFmW9JLlr7T0VI6X4j7Y0lN5t9WdHs5/2aWn97imO+ZDlp621XjrFoVxrNegClaAS2QFuMMb1HQFuO03qLgCXFpoULagQRWs5ZYur2TutzmnPYPqNaYH2pxTsbsRyP0Lb865xzXdMR/WmNzdGhPKr15rcgWtCVmybfDWNOZd2Ieza01xybZ0BwGuXGuqWD7KWrWTqsj8JBb9DlOM2Bw7FgaPU5L4xHr0xiQm1gnbnLAtCNuSsNXRZmNoG+qglvwYqulwr+sODttDmZrg2ZBOxdGQDm46w01z3LTATUvcVCdNw8sUBU9GbSa/cs/Z++XV32YhDBlbu4zP0mu/47P0Bs/b5bd8Np9UTNMLMlp3k89TRU1z3LTATUvcVLcm4jF7lcwHear2KSNUgEaHaQPJmLIns6PZAHP60aSnsXjiIXyM2/5aLToeRAWD4rnwdWNdWclgdarnj2VSMaKYhogRZWNCBeepYxpC+4e1fLfrZYeMqQoY00xuYfFXaKWP+Wy/UuyLcuZRkhTt9J3rqPvHMxhaKD26W7fpqOUdc9SbXv20Tx9OhVHz2eudt3HoV1XtAhL46wCuJtPRqwJM6UduBEz3Fy/+/igDbPdYhx2w0JysGwOrCoDNP9G4TZIGNpu93nnbAas6YBUAO32Q4RBYXQAs/fjeBtgfn7/JAasLGZtfkDRHXeoBqwlgs9nrnbcdsLoDVgOwmdUmlSkAln4cggdWzPatdV+gL0pjaqZkFVJYNcbUFGCKedPD1BCYZrPXO287TE2HqQFMTQZTW4CpzZJVzvYBzy+AaxpTm+CpghHTGFNbgCnmTQ9TS2CazV7vvO0wtR2mFjC1GUwLniNc0du8A6Ztz0oC6xI9q7Jm0gG4AmAxl3rAOgLYbPZ6520HrOuABaVYZbSJKlgdpFierG3PSgHbFpNlbJuOBBZ1aQdsmyQJbD57vfM2AqtYBFaxA7iaDLC8AFh6J2sA9lv/fzXb55oxHFo+5aw21o0526ajocWc6kHLCWiz2Ws13uEJBXbQcoCWZ6AVBdCKLGc9qEwbHFQx5aviXPAxqKIAVMydHqiCADWbvd5524EqOlAFgCoyoBYoSEU/cDOAugJUKyFwVOWUqhU3qhqjKgtQxfzpoSoJVLPZ6523HaqyQ1UCqpkF4qpAZam8ylJNz8pZxdOgpmSWkWzStRbILNSdHqiEzMpnr9VEZqlOZimQWSojs1SBzFJ5maXb9o+jmtBYSjpjx6gWaCzUnx6qhMbKZ6/VRGOpTmMp0Fgqo7FUgcZSeY2lYweAw5pQWFXlrBjDWqCwUId6sBIKK5+9VhOFpTqFpUBhqYzCUgUKS+UVFudtF8BdlUY1pbGs5BNUCzQW6k8PVUJj5bPXaqKxVKexFGgsldFYqkBjqbzG4rIdWOGwJmSWYU5PRlYFMgt1qAcrIbPy2Ws1kVmqk1kKZJbKyCxVILNUXmZx+Svg6pQicE2oLGOYmtC1QGWhHvVwJVRWPnutJipLdSpLgcpSGZWlC1SWzqmsN2cAqmLqi2A8PbbSCYVlmdVjUHWBwkLd2YGqCYWVz17ricLSncLSoLB0RmHpAoWlcworgGo5I0BNaCurtBmPrXSBtkLd6YFKaKt89lpPtJXutJUGbaUz2koXaCud01YeVFFxR4Ca0FZcWz2ZZtUF4gr1p4cqIa7y2Ws9EVe6E1caxJXOiCtdIK40La5e/bx4zA237QwW1w56V8t1Gt6EyNKSy7HI0gUiC/Wrhy4hsvLZaz0RWboTWRpEls6ILF0gsnRWZLXAammAtoylcU3pLCDthLUFOgv1qIcrobPy2Ws90Vm601kadJbO6CxdoLN0Vme1uBqrCFwTSks4ycdKSxcoLdSjHq6E0spnr/VEaelOaWlQWjqjtHSB0tJZpRVxdYzANRXMSuFaILVQj3q4ElIrn73WE6mlO6mlQWrpjNTSBVJLZ6VWi6sV1RfuMFxTYiuFa4HYQj3q4UqIrXz2Wk/Elu7ElgaxpTNiSxeILZ0VWz1c8WFBQm3JSvsn8A9xLVBbqEc9XAm1lc9e64na0p3a0qC2dEZt6QK1pbNqK+JaKYKvCbUlHKQe41qgtlCPergSaiufvdYTtaU7taVBbemM2jIFastkY1oRVyXw/tUkBJfkyskRrqZAcKEe7XA1hODKZ6/NRHCZTnAZEFwmI7hMgeAy2ZBWi6ujhKxJaC4J46xx/2oKNBfqUQ9XQnPls9dmorlMp7kMaC6T0VymQHOZbDwr4kqNX01CdsnKTHEtUF2oRz1cCdWVz16bieoyneoyoLpMRnWZAtVlsiGtKLgYs2hQyyT0loAc49VYpkBvoR71cCX0Vj57bSZ6y3R6y4DeMhm9ZQr0linVWyEEA3f4NK4JvaXhvjWeezEFegv1qIcrobfy2Wsz0Vum01sG9JbJ6C1ToLdMqd4iQ9smpbdSfC3QW6hHPVwJvZXPXpuJ3jKd3jKgt0xGb5kCvWVK9VbAFe9gE4JLwYBgQtgCwYW61AOWEFz57LWZCC7TCS4DgstkBJcpEFymVHA1cRglEGATiotry8UE2QLJhfrUQ5aQXPnstZlILtNJLgOSy2QklymQXKZUcgFO7ItWCLAJyaWYMOOpQlMguVCPergSkiufvTYTyWU6yWVAcpmM5DIFksuUSq4wxc0V1hWk1hEyU9nx+iFTILpQn3rIEqIrn702E9FlOtFlQHSZjOiycbsZsfvHtvoF2YGCbTcb7ga6+/6yePrUkzkJ25ywLQjbkrDV0dbA3e4vq6pjqIfB/rLmUAb6/P4yi+8vw01nuGmOmxa4aYmb6qRpeJkF+8ts2f6yH1YfVuez3UP/gDtKDXaSTV/jgHAK3yGGm+a4aYGblripttkdYrZADNmH3SEWi/8zdojFc33NHWKxTGqHWExznx1iBeepY5oH3CFmC2SepWXeD7885kkmJHRdxZwdx9Fsga5DXehVCqHr8tlrO9F1ttN1FnSdzeg6W6DrLK3rPJCz/ZeLx+kNNjah56rKTOYhbYGeQz3p4UnouXz22k70nO30nAU9ZzN6zhboOUvrOcBTJoFMrU00QkyALNBvqAs9IAn9ls9e24l+s51+s6DfbEa/2QL9Zmn95oEMb2Fo+lApjU73/Qn9Zlylx4NhWyDfUJd6wBLyLZ+9thP5Zjv5ZkG+2Yx8swXyzdLy7YdfbqDJa6H8KAVug2lUU/EyLSbzY7ZAvKH+9FAlxFs+e20n4s124s2CeLMZ8WYLxJulxVuDqsmgmlBuUis54WqBcEP96aFKCLd89tpOhJvthJsF4WYz6sEVRMscHS1rULU0qi4VK9NmMtHgCmJlqD87VB0RK8tnr3feRlRdFytz7ACuJoNqQazM0bGyBtXmpp/rCFwqYOaUGAd2XEHADHWrBy4RMMtnr90kYOa6gJnjAG4mYOYKAmaODpgNwM30By4VNXN2slbZFUTNULd64BJRs3z22k2iZk504AoANxM1cwVC0dFRswG4uW4htR+MMTfeG+4KQmeoWz1widBZPnvtJqEzJztwJYCbCZ25Ak3lcprq5rHMUjYlsMQ00OsKBBbqTw9VQmDls9duIrBcJ7AcCCyXEViuQGC5nMDyqOa4mpJZkvHxwMAVyCzUnx6qhMzKZ6/dRGa5TmY5kFkuI7NcgcxyGZnlHxBjnF8ILqokogm9JZmeLKRzBXoL9aWHKKG38tlrN9FbrtNbDvSWy+gtV6C3XEZvAaKPOQug+hf9JUFNrU4U2o5Xe7kCrYW60wOV0Fr57LWbaC3XaS0HWstltJYr0Fouo7X8E0w4txznaUJoVU5P9oK7AqGFOtODlBBa+ey1mwgt1wktB0LLZYSWKxBaLiO0AFLP04AqStSk0GJuIgkKhBbqTw9VQmjls9duIrRcJ7QcCC2XEVqAQh7WkIimqpxxpRlK1VjCcNEM55OHQcSEJKy4PztcY5oksAUF1D2Xu2eusk5uwZ8H/qpy6BYorpCIZq0MrPUAY6yNhQyHAUqKcf8aE2YAzsutmAYBOC+4ei7vAOY7gLkHOCO6OCtQXSERRd863L9MGteEzqqsmWxfjAkzuOaVVkyD4JrXWj2Xd7iKHa7C45rRW5wVvAA8JLq5fegcC3TePYjeeZKKolPGOWVcUMYlZaw74yCUbqtjXzuDWHp7LFcdVTaaHtLcpMPphO2MsM0J24KwLQlbnbaNLleVsK8VNGiwsX1s6/zl27/Pdm+XaZZspFnUvlAjiRJqmxO2BWFbErY62ojgOWcFWigkunm48HlX/p8RP+9O9jUD6F2hVAS9S3SfEHrJmeou0QMG0Tkr0HwhEXHXbNrVcvlydvo9E7N9pb6EhW4pjpjpPZRzqdR4Pi2mzNxEMcf69UXov4IC6p7Pu5uo2d1Ejb+JZkQgZwUqMCQiRn87mF99zyq/ZfWLw2DuXn3Qg1nqSk5hLlCEuGN9mAlNWFBA3fN5B7PdwWw9zBlhyFmBMgyJsmxevJmdMt5wOb2AMpYzfPaK42Y6zi7Qh7hXfYwJhVhQQN1zeYex22HsPMY5mchLZCJnBVQGjF8JFYgcnv6UwLgtZ/h0C5PoLniJWES96mHMKbGYL6DuudxhzHdikXuxyHNikZeIRc7LePzK98mAMdYnt+UM+2Tm/01ALhGMqFt9kCnBmC+g7vm8A3knGLkXjDwnGHmJYOT060Z7INM9MheJHplpkQC5RD2ibvVBptRjvoC65/MO5J165F498pxc4QXhupCojMk2gMwrpEtuCxpSWXDoMSYoF8TrcL/6KBMRu4IC6p7PO5TlDmXpUc6E7TgviNuFRFkq/9JSeayS4Nvq5lXQXDB1guBfTVk+JXhBUA93tg89EdYrKKDuubuDvtpBX3noM7E9zguCeyFRluBeOFjWvPZUct2+USS+EaR9owgM79PIqynzlZjsoI/pMuBj7vbBJ6J/BQXUPY934Ksd+MqDnwkBcq4LZge4vtPcVFcpd5+LimdOzkURxjllXFDGJWWsO+NgLgrUy7GvjsFkVDyYqwCTn43iBp+Nwm1nhG1O2BaEbUnY6rRtdLm2hG+2aDbK97AizK0Ak3Tzqiqtq/ANXW34ZsIMXjAkbdMdSNcwkFuBMM/i81e4bU7YFoRtSdjqaKPmr3iJbuMuN38lrePN/FVVpdf5doXgk1R+1qa53znrbjcp1RaeRha1LQjbMtp0Eln38BNEokTuiezjSm5B997EnbV0BysmYu2PZ46x6W0vJqxSVUMYF5RxSRnraGxvcN5Zui8VJaJP0BHCV2+/fzvbQSo0Yxh0g5CbUA1qqVcVxqQ6yWvCuKCMS8pYj9yD29T+6Y9vf/g5vHeQH4b3Dj76dndITg+p6SE9PWSmh+z0kGsPPfJ3R3bo0Uu8KHFUnSXyUtDxyGbeXQ9fFsdE+9K45sWZzegk3iP8Fw+Nybgwta64Di/UbOekG0bs0kNXqdVJ/zz+lexpxkz05B/PtEo1NkE1Nty4oIxLylhH466xiUztlOhSQa8jbWpDDmqjUqNXAIYurOna0qi2+hF5x2UikCGFMbtAhmQsvRDgrCs6XQ+4cUEZl5Sx7oyMuNuLqFYdhX2rEi3VNLQZNI3KRgqzgvvLbd/HaJWrmhZkm761Ci2vfV1vvF8lX9obr+Ye7zmfEjrWcomcxtHs1/pET//xbLSCbBkTdW+O3ff5ei8g9j/Rl8iOcPAvjoVD7Ztj4a8Dj1SuXy2JCAt1J82HjbHvpQKjL0kVSBjnlHFBGZeUse6MnQrc94OV47oAe52Xe0Ljcg+3nRG2OWFbELYlYavTttHlmhKqmVK51zZ1obWtelSScIfGqGNwGYfb5oRtQdiWhK2ONkrGiVYeqxRco6St4vMdAg6tI1txWLFQMdPrRsfNFMYBrYaQMOyZLb7vvQg6jIRmoenCkIl6a/QuUZwO7R8ZLl7qW7DOpLsDOWdUrxPp3RRYs06CxWEFN5lOxlGdDG6cU8YFZVxSxrozmtDJvIah9Nn+fi3Ut3XFvq0tfDiT3/opqW99j/ToyK9nZHDnOOApKTJkj2TFRJMlqkrSqsrXDJOyGS6H27RwygxWrNzqdn6LRSmn0beUTDqjjHPKuKCMS8pYd0Z797GMzFSvKOgcZDPil/gKIP/8ztC8nAmaSVTQPfhv2Tw2ZefvoJPI9SqhyZuT2T6Hv6wRTeeeJEg6StJ6rl2yRnHjnDIuKOOSMtbRaFivrfoWfLTPH/v6vmXjlOWNsyRuJem4FVHPqmqWmf2JjbWiGitunFPGBWVcUsa6M96jsVaZ6lYljbXdD4bdyU9/KmmWRe053TyLKII019b35K2VMM4p44IyLilj3RnloLlK0TbX6pbNVZc315LFgpLeIRZq+n+hYRqqYeLGOWVcUMYlZaw74z0apspUbAxWkbXVrszD5ct0BjI+fExx26sM7x/rVw7dnKctPtFwp2RBmmh7FenaxY1zyrigjEvKWHdGPWyiqm2i5pZN1OVlt3S47MZtZ4RtTtgWhG1J2Oq0rbnco5uP6/X2bLVdPXtyub7+sD5dX1zczH7b/P4J0ko/OukOz67X7xv1fRyk6VHCxu1xiD5Obafq+Ewl80hxHAZLyfL0cYivp2xMHod9IMkyzXHoDBK26riuksfZ8bJiKYsFi01aOFPHYbdAwib08ZlIee7DUschkJXKdVynkPCT1sdhHjh5teo43L9SVwUFVukSuTsOsc6UF1BZAqktexwaYsomfIkiWSLcI4/DuCmF73GdZAx0X8chBpE8l79mxEeok7QP/DgIrZQPgJNNngnofpZmO1xu3Vzt0a7tPHvyDlrTL6uLc/g+33zqGpPaG5tmN/8dClr6wcXHzeez683V2ebzp6d7rDnw4tPV79tX65ub1Yd1d3B+fb257h9cXVxsPj+/WH36R/i5/ecVHL84v9n6jnBzffn7xYo/82svg6o64U4cQkV0P/agI4ypnhwNPcQ8PtVfzWOOevxi/pir00P4Etz/L4X/vzoLP5QO/zv/vw4GLcP/Vfhfhf/1Xa5twc1sEcL8X+0C197+8/n2Aqxh6kkGqct51QRTurt9DILA3bzNFTKEFQX73FaPmompZoJKtvEu2dyrhRXyZB/SNolkE4dxYd7Tx8qaU0ij2kzNwEG28Rxp20LbgQ+cEly4ut5cXm07z9M7k8anidl8DgndUrAY6fTJ02YI086v2TaSJOTwdxchGtvtCcqVAUiHAzTuxO8QU3voJvl6tV3zQVyjaiZCrdvFNw5xzPzqjjB81dKYk7tc55nkszMp4aPgo+Fj4GPh8/X4z+7J/8WLHzDWh/njB+L6ZFqaWWYJroObT28f4Gwmuw/9TPjTaZwzPTWO8glcOPRrJkgijA7A/eoKau3V6vrDOdyrLtbv4VrYX/0Q+7oZ8DY/tpur0JX9utnCYDj8+XG9ere+9gnA/n6z2cYfR02Zb9bb369m78+3P2/+dv5u+9Fn8vfKz5vrf4Rh57P/AVBLAwQUAAAACAARTWhcDDdeB4wiAACX3wAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQ0LnhtbMV9f3PctpLtV5mnV3VXdpQVARC/JFtVN9LMMJs4SdlJ7lZt7R+TeGyrruTRjiax73761wQJDkigG9DIyqvUZKxpAGyePgBx0CD54tNm+8/7D+v1bvb59ubj/cujD7vd3dnp6f3vH9a3q/t/39ytP4Ll3WZ7u9rBn9v3p/d32/Xqrat0e3PKq0qd3q6uPx5dvHC//bS9eLH5Y3dz/XH903Z2/8ft7Wr7r2/WN5tPL4/Ykf/h9fX7D7v2h9OLF3er9+s3690vd1D+3fXu581P8ENvOx3afHt9u/54f735ONuu3708+oadNUzYtowr8uv1+tN98O/Z/YfNp+X2+u334AicV3U0a8/1t83mn63527ftT+2xP65nn9/c3Vw7b2b/6v8pjma7zd3363e7y/XNDRyuPpqtft9d/7n+CWq8PPpts9ttbt1ZwDntVjv47d1287/rj86j9c0aCoOvd640NNUVTdi6ltoj4db+OJ0DnUN/b7H8HwfF3wOk2nML/+0hWbgAAsC/re7Xl5ubf1y/3X14eWSOZm/X71Z/3Oxebz416z4osm3v983Nvfv/7FNXFiD5/Y978KavCw7cXn/svlef+3iF5TlSgfcV+KQCq5EKoq8gphUwl+q+Ql16BNlXkKVHUH0FVVpB9xX0tIJEKpi+gpnCipS3fXnXJU676LnQX612q4sX282n2daVbkMsKt/KEHToDL+3Jb5pieUKwq/XH9t+/Ga3Bes1NLi7+Pnv3//4t//LuOTyvP2uVd192/5v2f+txfmsK8C6H4zovw18CyYq676VsO536b55JUzbEFfKtA0JKYQays1m/xVX5ZURuv1mrG6/leTuEK0vdvCl++4PxXhfzp7/94vTHeDVntzp7/ABnAaweAcWUzhY3IHFEbDag4mqPY+6UrU7HyFsB5Co3HlKqxweWnCdOBmuhVUtkOAtIDzbt8i1rqpUi6yu29PjthKsP90+JEZ3f/MODqF6WFwoheDSuPrct6NcO3WtWQAr59oKFyEhVVfO6AHOhKOiZsY5yqtaoI7Of3ndoeBCayy0FZ41Zx0PPD+gEeuDWaWDXPvf2fjvnhzCso5EtSOR1RU/J7ggOi4QVBCOCsJRob0atk04yyVquUItc9SyQC1L1NKkLKPTq7vT48S4ULs2apTqVgnexV1W4nw0Rmje90c/ZoRDAZhd6DiXWk9C0OHXHVkm8EMtc9SyQC1L1NJ0FoXjJ/P4SdeGJoaKB+LGJEvj1R3JJPCSifPo8MK82+NW4P/0mqD68Vf256HrlL/L0N+4Zca4O9lzZnmidpM621FsVD42Kn9uyUG0i8FDrzmpmCk0ZgqNGeb1Pma58xISLjTB5ZirCobXVIy6lliV6Bwp10cB0NmxU6NjJ2q5Qi1z1LJALUvU0qQso9MzeX6ZB4yd/XRJ9pctfxHn479r0/MsOb3iMNvovqvKDvx011hei57KwlYkJQ067KKWOWpZoJYlamlMbti1/QzN4NDbjrjYfHaKVQqIvgl8ljeencwA50QzV30zCmnmODV3G+Z2Wp4/SzQ67xvFqRVNsxKtLPpWJN7KdF46mremBozM2TqwHMe5tZaXNNn0TeLDGXipq25qWekhHsSsrh3PHF0soYeqB174YIbM9zPm7zfv/7G6ufEd7b9u+X+nONYfBZh88eLdxbeL42/f/PDLq2/mr4/fzP/++rI5BmlZH50s5bNnJ/+Wmg9T8+B/+z+X6qBaEjj37uLFn+DpnyGVUUz2xOyL8MqdUJVsZ+ELscSw4G3cIwKl/3azO//b+9350dEJ/HVSpd1rfFXhq0JjL6sTqAX/eA5YPz8EDBEdbcwlVsAlluESiLGKdwM4CB+28Cxa3Gw22282q+3bmf+R4BILuXTIqWok7Jj7QdhZGHaA/kowCPS3PxydoPFa+EpJGrCIBmxEA4bTgE1osPfFcYJ5TjDgBDuQE9BChhe8gBf8QbzgKV7wPC/4U/ECcz/gBT+EF5zgBY94wUe84DgvOM0L7nnBgRf8yXghCnghcrzgxl17eMWcEJmOG5fra6j0vl20vikZOcRjGWIQhmAnEjBERAwReYYIgiEiYogYMUTgDBERQ0TIEOEZIoAh4skYUhcwpD6IITzNkIIxpH4qhmAnEjCkPoQhNcGQOmJIPWJIjTOkphlSe4bUwJD6yRgiCxiSW7gBd6wTQOCPm78qLcV+/rq6vf642q3ffrNe3e7ZIdLskI9lh0XYkV2fmvdFMhNOSbBBRmyQIzZInA0ymnBKTwAJBJBPMuFUBcHPr2zVtlvRqq1ffQyCv+kGhuyooB4bd5YO2BV6AkHgVTQsyPywoAgiqIgIakQEhRNBRcOCDIcF5VmhgBXqyYYFXcAMnb1w1H7KWZmBEa83m3evV+92622WE3qiaBdMByh2P7w86q5Q4vyYC/ns6CCVClO9g6qxw6Q01EO42p0xr/sznr9+/ePr4yv8KjL3NQLq2oEuXXXA8Tm0DE3gXO5bSS2ueRsf+bSkrmw6orANKaw9hfXz1renorApoLAhKdxOd9s8sltkAo6pgcRsMXPC6T821/e7LI/No8c2gfDFpPmCKtq5rxHwRUd8McAXQ/KlbyVemwt4Y9K8IdS2iXijQ94YzxvzvPXxqXhjC3hjSd5wlDf8Qbyx0fhnp+Of/TLj32GLe0weOP6hq4I2zWdUic99jYDPJuKzBT5bks+WGP9smsfE6oCNeGxCHlvPY/u89e2JeMwLVqZ5bmXar0TzlqkDi9vl6MV2dbvOMZhXjx75kAWjvuWIKejEas6riClqyhRePYeWKab4VqiRj1dpxuCTPj5dbg59O4YGe8aAf62PT8WYgvVnnlt/bhOI6jxMOHYOdcz5eb3drn7P8+bR688MkYGo/wFRRgvQiAzkxIIzjxac+WjBmeOXQD5dcIbGfOwZxP7QNWY67gXry5xeXwaqdqH3+zWq8yBn/OPsCq55Q9CdImB8unOuD/2jl5g5IgTRUwhCz0tCT6wp82hNmY/WlDl+1eDTNWVozIeeQ+gPXUamQy/yWw94v5RJ9XlhfLbZrQNVzPDz16vd2iXFa3cFqYWQ57NZu1VBJnfxXPoDYft42gZZsI+wMm7nnVCG3odw5Ru2h3OKIReVfl9cnNkP+JIvsywo09BlxmEtWPfl3fqjEA8M6z4Afu8Pr2y/lbJ2U12/H0Iwqdh50O2Fsm5PLlwaauX4kR4Ces+6aztED+nPfTF8z4GRlp/PUjzsTsr9rXWX8RftoJXaIOGPo0aX8ze/vDqes+pkDoJmDpPBORPwqeEj4aPgo+Fj4GNPYO4BHyjH+TPoyfXpMfsavohZRl0wy8iXaXDXG3C9AdcbcL0B1xtwvQHXG3AdBDx8DHzsCUxN4APl8q6PGej3M3KCgd2ap0AGlvaGkrP7O5guvDy6267v19s/10cXs9niu3h7UruNjp/vd3sPFJ1s9cb3KnlnUhvoCNvc27qxpWfFGZAGC21Xvq6o0ObLNPFxIaRnTeK447iogrh0S5A1suOJjMtk/1GfY/5+uqOGjER/eJ6KBG6be1u8kylAP19mWVCmocuMEdcFl9huxayW/tpP7N05aq+hvK5qd1dDVbkdQ2m+Q0GjOQyBi++gfjo69vw0bT469Ejdblv0SLE5liU9EXpUEjtXrwjb3NvieURAhHyZZUGZhi4zJoLJ7mLl3QpYahsrbrrCTXPctMBNS9zUJE3j07QFfO9WSGqb3c/a7Sf2iqEd4CXrJn22DvepdrT8j5+cDPF3AvV7+Yr3q/ZuycQe5SvCNidsC8K2JGxN2ja+o6TKb14V3cqCfMzu1YI2JvPpYIoOITi95dTs3DeP7Y4FbTk7nUFoU5OzvnJil+g+BL7Ml9ycih/38N2pvk1MaDl2P3iHqihY1REMG8W+6LbC/jBSHK7DkL0fvmVsFzNwKMke9LwD9rD8XFxMVnzyuY4xFH4HGUS4W/ETftVHsOcAW37tThSs4Yi/ZI+gGK1jfNEgdy3LbpYEgKU1OXqaQUx5QUx5Oqb4Co5I7grcx9Qv5wgOMeUFMS3Y3ydEtud+2Q1+/QEf04eRpVnf8kP7MIpAEG9REG+Rjje+x28Mhd/BtY+33+MnBMS7YP1dFKzaiL94t56oH92bsXDX496cXuWZoyccRLdgtUTU6ejiuxhEcn/ePrp+f56oIbp1QXQLdtoJmevNX2a3VX+cR3RigSyy+5Yf2onREw/CLAvCLNNhxlNvYyj8fqt9mP0uPCEhzLIgzAV76kRuT90j065CPbbTComEVx0W3vyNpt5nMroqGV2O38cj4t10Koyu300nFERXFUS3YF+cQJcEvuiuov4wj+nDCgmyPizIBcsufRk6ysiGM4NHWUd9WIdR9hvOhIYo64IoF2wdE/TWsS+0BUhMt0Q92RYgYQ6rpg/bAgT1EPKZw8iHhSPgXsGmNYFsWrM49+JNaybknt+0JgxwzxRwr2D7mbD5Sf/jtt72R5AB755o621dHcYfeyDvLMY7exjv0FAExLMFxEN2mWmceNMQuV22e+L5XWbCPndo5YhXV/nMUN0v1SGrXlhmyC1cYSu3WOrOL3Thy7XeG51YriVsc8K2IGxLwtZ4m/HpOIhBI9gZhOnkaOg5J/1PmUiwbKKgZmiiADdd4aY5blrgpiVuapKm8WnygkcGdUsquede/PrqHzOXkvKdXfhnW/Tf/tkWowdRPfRZGUJWik4i9+4mHzSEmua4aYGblrip6U3EYy9qkU8d1P2tnNSyv3JiVFSVNOez09kIc/qZPZe+eeKhGBUz4f4SOstApRj8sfC9LkNbycRlauT3bVKZB1+GyDxkMw0Fx2l8GUJRCqn0A0fZMWPqAsZ0SyZYVg966ddsdlzL6rO0OvUwksu+BWkH6v55wZmRanK17stROfo56k0Qn/6xXKnkXL56s/fWz5Lq+iVcc9udfnX9HM4mM9DLAkzpmy8dpseLb//zWQbY4f7CPbDQnYydAisLgM0/6qsvkgY2W73ZezsAKwdgJQAb30AwBlYVAKvyZD3+8Zs3OWBVIWPzu0rmqEsBsIoANlu92Xs7AKsGYBUAqzLA6gJg6fsCW2D57NgY+xnGojSmOiYrF9zIKaa6AFPMmwBTTWCard7svR0w1QOmGjCNpe4YU1OAqcmSVcyOAc/PgGsaU5PgqYQZ0xRTU4Ap5k2AqSEwzVZv9t4OmJoBU1DStclgWvBcr5q+icth2o+sJLA2MbJKo6MBwBYAi7kUAGsJYLPVm723A7B2ABaUYp3RJrJgz4ms8mTtR1YK2L6ZLGP7ciSwqEt7YPsiSWDz1Zu9tx5YWXlgZfUcziYDLCsAlr75xgH7Vfv/enbMVFXh0LKYs0obO+VsX46GNntHzaIvkoY2W72R05tioMEBWgbQxrdTj6HlBdDyLGdbUCulcVB5zFfJWHvrwBhUXgAq5k4AKidAzVZv9t4OoPIBVA6g8gyoBQpS0g8fcqCuANWacxxVEVO1ZlrWU1RFAarZZwgt+iJpVLPVm723A6piQFUAqpn7cWSBypJ5lSW7kZVVNUuDmpJZWlTR0Fogs1B3AlAJmZWv3shIZslBZkmQWTIjs2SBzJJ5maX6/o+jmtBYUlhtpqgWaCzUnwBVQmPlqzcy0lhy0FgSNJbMaCxZoLFkXmMpPwDgsCYUVl1bw6ewFigs1KEAVkJh5as3MlJYclBYEhSWzCgsWaCwZF5hMdYPAczWaVRTGssIFqFaoLFQfwJUCY2Vr97ISGPJQWNJ0Fgyo7FkgcaSeY3FRD+xwmFNyCxdWRXNrApkFupQACshs/LVGxnJLDnILAkyS2ZkliyQWTIvs5j4DXC1UhK4JlSW1pWM6FqgslCPAlwJlZWv3shIZclBZUlQWTKjslSBylI5lfXmCkCVlfzMK5aeW6mEwjKVUVNQVYHCQt3Zg6oIhZWv3qhIYalBYSlQWCqjsFSBwlI5heVANawiQE1oKyOVns6tVIG2Qt0JQCW0Vb56oyJtpQZtpUBbqYy2UgXaSuW0VQsqr5klQE1oK6aMipZZVYG4Qv0JUCXEVb56oyJxpQZxpUBcqYy4UgXiStHi6tXPi6+ZZqZfwWLKwuhqmErDmxBZSjAxFVmqQGShfgXoEiIrX71RkchSg8hSILJURmSpApGlsiKrB1YJDbStqjSuKZ0FpI1YW6CzUI8CXAmdla/eqEhnqUFnKdBZKqOzVIHOUlmd1eOqjSRwTSgtbgWbKi1VoLRQjwJcCaWVr96oSGmpQWkpUFoqo7RUgdJSWaXlcbUVgWsqmZXCtUBqoR4FuBJSK1+9UZHUUoPUUiC1VEZqqQKppbJSq8fV8PozsxiuKbGVwrVAbKEeBbgSYitfvVGR2FKD2FIgtlRGbKkCsaWyYivAFZ8WJNSWqFX7KNgxrgVqC/UowJVQW/nqjYrUlhrUlgK1pTJqSxWoLZVVWx7XWhJ8TagtbqH0FNcCtYV6FOBKqK189UZFaksNakuB2lIZtaUL1JbO5rQ8rpLj46tOCC7BpBUTXHWB4EI92uOqCcGVr97oSHDpQXBpEFw6I7h0geDS2ZRWj6ulhKxOaC4B86zp+KoLNBfqUYArobny1RsdaS49aC4NmktnNJcu0Fw6m8/yuFLzV52QXaLWMa4Fqgv1KMCVUF356o2OVJceVJcG1aUzqksXqC6dTWl5wVVVBk1q6YTe4lBjuhtLF+gt1KMAV0Jv5as3OtJbetBbGvSWzugtXaC3dKnecikYuMKncU3oLQXXrenaiy7QW6hHAa6E3spXb3Skt/SgtzToLZ3RW7pAb+lSvUWmtnVKb6X4WqC3UI8CXAm9la/e6Ehv6UFvadBbOqO3dIHe0qV6y+GKD7AJwSVhQhARtkBwoS4FwBKCK1+90ZHg0oPg0iC4dEZw6QLBpUsFV5eHkRwBNqG4mDKMR8gWSC7UpwBZQnLlqzc6klx6kFwaJJfOSC5dILl0qeQCnKrPSiLAJiSXrLieLhXqAsmFehTgSkiufPVGR5JLD5JLg+TSGcmlCySXLpVcbombSWwoSO0jrHRtpvuHdIHoQn0KkCVEV756oyPRpQfRpUF06YzoMv52M+qlt71+eeCDCMd3Ax1+f5k/fOpBhIRtTtgWhG1J2Bpv6+Du7y+r6zOIw+j+su6nDPT5+8sMfn8ZbrrCTXPctMBNS9zUJE3j0yy4v8yU3V/2w+r96nq2f5QccEfK0Z1k8ZOvEU7hd4jhpjluWuCmJW5qTPYOMVMghszT3iHmm/8r7hDzx/qSd4j5Nqk7xHyZx9whVnCcxpd5wjvETIHMM7TM++HXr1mSCQldV1fWTPNopkDXoS4EQSF0Xb56YyJdZwZdZ0DXmYyuMwW6ztC6rgVydvz94uv0DTYmoefqWkfrkKZAz6GeBHgSei5fvTGRnjODnjOg50xGz5kCPWdoPQd4iiSQqb2JmvMIyAL9hroQAEnot3z1xkT6zQz6zYB+Mxn9Zgr0m6H1Wwuke3J8N4YKoVV67E/oN21rNZ0MmwL5hroUAEvIt3z1xkTyzQzyzYB8Mxn5Zgrkm6Hl2w+/3kOXV1y2sxS4DKZRTeXLFI/Wx0yBeEP9CVAlxFu+emMi8WYG8WZAvJmMeDMF4s3Q4q1DVWdQTSg3oaSIuFog3FB/AlQJ4Zav3phIuJlBuBkQbiajHmxBtszS2bIOVUOjalO5MqWjhQZbkCtD/dmjaolcWb56s/fWo2qHXJmtnsPZZFAtyJVZOlfWodpd9HMDgU0lzKzk08SOLUiYoW4F4BIJs3z1xkYJMzskzCwDcDMJM1uQMLN0wmwEbmY8sKmsmTXRXmVbkDVD3QrAJbJm+eqNjbJmlg/gcgA3kzWzBULR0lmzEbi5YSF1P1hV2em94bYgdYa6FYBLpM7y1Rsbpc6sGMAVAG4mdWYLNJXNaar7r0WWsimBxeNEry0QWKg/AaqEwMpXb2wksOwgsCwILJsRWLZAYNmcwGpRzXE1JbNExaYTA1sgs1B/AlQJmZWv3thIZtlBZlmQWTYjs2yBzLIZmdU+IEbbdiM4r5OIJvSWqFS0kc4W6C3UlwBRQm/lqzc20lt20FsW9JbN6C1boLdsRm8Bol+zyoHKqnYDVwLU1O5Ersx0t5ct0FqoOwGohNbKV29spLXsoLUsaC2b0Vq2QGvZjNZqn2DCmGE4TxNCq7YquhfcFggt1JkAUkJo5as3NhJadhBaFoSWzQgtWyC0bEZoAaQtTx2qKFGTQquykSQoEFqoPwGqhNDKV29sJLTsILQsCC2bEVqAQh5WV4imqpgxqSqUqr6F8aYZxqKHQfiCJKy4P3tcfZkksAUNNIHLw+NJq0FuwT+ft2eVQ7dAcblCNGuFY20LMMZa38h4GiAFn46vvmAG4Lzc8mUQgPOCK3B5DzDbA8xagDOii1UFqssVoujbuOuXTuOa0Fm10dHti75gBte80vJlEFzzWitweY8r3+PKW1wzeotVBe9sdYXuH546xxKdhyfRB09SWXTKOKeMC8q4pIzNYByl0k191kZnlEvvf8uFo85m012Z+3Q6nbBdEbY5YVsQtiVha9K2yenKEvb1ggZNNvaPbZ1//8t/zvbvLOm2bKRZ1L+mIYkSapsTtgVhWxK2xtuI5DmrCrSQK3T/dOnzof2/In8+HOxLJtCHRqkM+lDoMSn0kiM1Q6EnTKKzqkDzuULEVbPrV8vl97PL7yo+O5bys9voluKIjq+hjAkpp+tpvmTmIoo5FsaL0H8FDTSBz/uLqN5fRHV7Ec2IQFYVqEBXiJj97WF+9V1Vt7esfrYYzMNbAgKYhapFDHOBIsQdC2EmNGFBA03g8x5ms4fZtDBnhCGrCpShK5Rl8+LN7LJiHZfTGyh9O+Nnr1im43l2gT7EvQoxJhRiQQNN4PIeY7vH2LYY52QiK5GJrCqgMmD8iktHZPf0pwTGfTvjp1voxHDBSsQi6lWAMaPEYr6BJnB5wJjtxSJrxSLLiUVWIhYZK+Pxq3ZMBoyxMblvZzwmV+1/EcglghF1KwSZEoz5BprA5z3Ie8HIWsHIcoKRlQhGRr/EMgCZHpEZT4zIleIJkEvUI+pWCDKlHvMNNIHPe5D36pG16pHl5AorSNe5QmVMNg5kViNDct/QmMqcwYgRoVyQr8P9ClEmMnYFDTSBz3uUxR5l0aKcSdsxVpC3c4WyVP61p/JUJcG3Ud0Lhhmv5DmCfx2zPCZ4QVIPdzaEnkjrFTTQBO7uoa/30Nct9JncHmMFyT1XKEvwVjiYqnuZpmCqf6OIfyNI/0YRmN6nkZcx8yWP7qD35TLgY+6G4BPZv4IGmsDjPfhyD75swc+kABlTBasDTB20NjUE5fC1KH/k5FoUYZxTxgVlXFLGZjCO1qJAvZy14RgtRvkfcwHQ+dUopvHVKNx2RdjmhG1B2JaErUnbJqdrSvhmilaj2hGWu7UVYJLqXlWlVO2+Yah13xXXoxcMCdMNB8J2DGSGI8wz+PoVbpsTtgVhWxK2xtuo9StWotuYza1fCWNZt35V1+l9vkMj+CJVu2rTXe+ssQ9blOobTyOL2haEbeltKomsffoFIl4i93j2cSUPoHuwcGcMPcDySKz9eWGrKr7s+YJ1KjSEcUEZl5Sx8cb+Atc6S4+lvET0cTpD+OqX736Z7SHlqqow6EYpNy471FKvKvRFVZLXhHFBGZeUsZm4B5ep48sff/nh5+4t9yfuvYPPvtr/JOKfZPyTin/S8U8m/sn2Pz1rr47VSYte4kWJk3CWyEtO5yO7dXc1fllcxfuXxtXdN+PBNaL9Yq4zaeuW1iVT4ny/Jt0xYl8ehkolz8PjtC/6TjMm0pN/XiiZ6myc6my4cUEZl5Sx8cZ9Z+OZ6JToUk7vI+2iIUbRqOXkFYBuCOuGtjSqvX5E3nGZSGQIrvU+kSGqKr0R4GpoOh0H3LigjEvK2AzGirjac69WLYV9rxIN1TWUHnWN2ngKVwXXl4e+j9FIW3c9yHRja+16Xv9mW3+9Sr7f1p/NI96eHRPaR7lETuNohlGP9PSfF5MdZEtfaHhz7HFbL3hXb/sn+hLZCQ7ti2Php/7NsfCv5y1SuXG1JCPM5UGaD5tjP0oFel+SKpAwzinjgjIuKWMzGAcVeNxOVs6aAuxVXu5xhcs93HZF2OaEbUHYloStSdsmp6tLqKZL5V7f1dv3tNcBlQRcoTHqaFzG4bY5YVsQtiVha7yNknG8l8cyBdekaK/42gEBh9aSvdjtWKgrHQyj024K84BeQwiY9swW3wUvgnYzoZnrujBlot4avS/kl0PDX8abl0ILNpgMVyBrtQwGkeCiUHX7JCo/rWA6M8hYapDBjXPKuKCMS8rYDEbtBpnXMJW+Oj5uuPyqqauvGgMfVomv2iWpr9oR6dlpu5+xgivHc5aSImP2iKqYaKJEVQlaVbWRqYTopsvuMs2t1KMdKw+6nD9gU8ql9y0lk64o45wyLijjkjI2g9EcPpcRmfDygsFBdDN+ge8Aap/f6bqX1U4z8RqGh/ZbdI9N2fs7GiRyo4rr8vp8dszgX0bzbnBPEiSdJek9VzYZUdw4p4wLyrikjI036iroq20PPj1mX7fxfmDnFOWdsyRvJei8FRFnWXfbzP7CzlpTnRU3zinjgjIuKWMzGB/RWetMuGVJZ+3vB8Ou5Jc/lXTLov6c7p5FFEG6a+978tJKGOeUcUEZl5SxGYxi1F0F77tr/cDuqsq7a8lmQUHfIeYi/f+hY2qqY+LGOWVcUMYlZWwG4yM6pswE1ieryGj1O/Nw+RKvQPqHj0lmgmC0/lVhcOjuHPf4RMeNyYJ00f4s0tHFjXPKuKCMS8rYDEY17qKy76L6gV3U5mW3sLjsxm1XhG1O2BaEbUnYmrStO93T+w/r9e5qtVtdvLhdb9+vL9c3N/ez3zd/fISyoj3Z4efZdv2ufXPA2ZU6Oo1+b1X5mZOsKRszZy4rGdsu5dmVTNYR/MxNopLtqTOXd0/ZKnHm7g9JtqnP3CCRsNVnTZ38vTpb1lXKYsBikhZWyTN3F0HCxgFAnvK8TVeduQRXqtZZk0KiXcw+c+vDybOVZ+66ljoraLBOt8jsmcuBpryAYHEkWubMddAkM9oWebJFuHaeuflUCt+zJskYGNbOXG4ieaz2nBEfISZpH9iZE2ApHwAnkzwS0P0qzXY43aY729N9n7p48RZ62a+rm2v4vt58HDpZfTQ1ze7/xzW0bCcdHzafrrabu6vNp48vj6ruh28/3v2xe7W+v1+9Xw8/zrfbzTb8cXVzs/n0zc3q4z/dn7t/3cHvN9f3u3aA3Gxv/7hZsYt2T6ZTW+fM8hMIxPDHEQyQvtSL07GHmMcLpmcLlyL/Mm7D3HXd2n++3t2A1S3bCCcTGau7RMRwpfQJBLgS9rVcBZeNP2amftYt6nSLO6LPFYnuOscNF+fHULYrJLochnVrhm2eqTuE0LKv1F10RZ8LEaZvtJ80wCHBhbvt5vZuN3ievqtnehhfra0hoOs6ixZWnb/sLv/92pTpszBcjP8esitTuzlHGTAC6WSExiEcuHT5qKem7evVbs1GOYG6W0Q0dp8bOMExa3dGuKmfElqfH3KeV4LNroSAj4SPgo+Gj4HPl+N/9Uj+L779AWO9W3t9Iq5HS7qVqQzBdXDz5cOTg91C8Um7ivwyzhGml5VRPoELJ+1+A5IIkx9gTL+DqL1abd9fw3h+s34H51L9ezs93XaTxe6P3ebODWW/bXYwkXT//LBevV1v2wJgf7fZ7Pwfp12bb9a7P+5m7653P2/+cf1296Gt1F5PPm22/3RTtov/B1BLAwQUAAAACAARTWhc38wFV6ULAACoQAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbL1ca08bSRb9KxYrrWbRyK73IwakAWJsyCaI7Gw+d6ABa2x3tu2EzL/ferVxdVd1FQQYKdNxux637jn1uKeuc/BQ1X+t78tyM/i5XKzWh3v3m823d6PR+vq+XBbrYfWtXKlvbqt6WWzUx/putP5Wl8WNqbRcjBAAbLQs5qu9owPz7rI+Oqi+bxbzVXlZD9bfl8ui/vu4XFQPh3twr3lxNb+73+gXo6ODb8Vd+bnc/PntslafRttWbubLcrWeV6tBXd4e7h3DdxcE6AqmxH/n5cN65++D9X31cFbPbz6ontVAwN5AD+5rVf2lv57d6FeqfLkorze60UI9fpQn5WJxuPeHtux/pps/jE2jbbu7f2/6mxh3qOF9LdblSbX4Mr/Z3B/uib3BTXlbfF9srqqHaemGSHV719Vibf4/eLBl8d7g+vt6Uy1dXWXAcr6yz+Kn88xOeTXycAXkKqBWBcgiFbCrgFsVMIpUIK4CaVeIlKeuPG1bFOuAuQqsXSE2Zu4q8HYFEqkgXAWR24N0FWRuD7olCxxoVUGxPuAWa0s5yxFDsNNiUxwd1NXDoDblNZHQFtAttRSfr3WJY01f5Q7lF/V6vtKT7/OmVl/PVYubo3/+A0GM2Vg9ARZUPSGiEJgn4dg9pX1P9HtGkRDmM+bQexJMdDuIS10PMUyZ+57b75F9YolMf8K0hyEG0jwZlq4/uNOf+tz0J7V9CBJq24eEuPadnciWI5Ds9KvtleOD0Ub5UY95dK3+KP9tnYisE6GIOxEZJ2pUIk5UnUo7OIqc8RD7g+BtI0zTJ65pFMcHUMhCVU8zqjKsrcCIK5f2uACnXYAtj+J9CUGFhZ01niABGBv6KPg5GI/e/3kVhsh6x/bKgOlWbyQ/jqCgB6Mfu35wpom4bY6KiAGMLHW4RymPsj1+Imk/kYifHsdkS4jtkB4HQmID2RaZuCIyMtZlez4HnHqWaMP4wUwdBKF9EtVmqKlpyhyUYc8saQ+mwDbBiYSmKQINh6hkog8vmsaLJniteWyXRcLc8qZm+ABCMzQ6hgLoLzWpFZdHSxSise0E6V5vj84Vf281k1s0pv00nnwwax/gu2u11A7QDlJMHg+OAcTaSxQbLxFAjNcIHwfNmtoehTe7OBhK37BZqJhf5HxbRA9wSvdnnTF6uLA0LuyZuGDyBFyYhwsL48LeHBcWcLgEQ9Ly+ixUrgUM84Bh+7POID1geBoYntoICAJ2kSVmcbUL/mRRVfVxVdQ3zXKfRod76PAwOjy1+OutWLS2YrDzWZ0+ADNYESbHPfvthHfdjYaI+AaduVJwpxQYqs1Oev/5laZ8F6YJH52FRzsLmNBG3GtqyvdnnaY8xEUacZFCHAkzFRFQ+0UH+ZNyrqrcXRarcpGPvfCwF2HsxdthLwKOHzLWwl6EsIcthKbCA1uMzsLDm4X6bIHtNTUV+7NOUx7YMg22TICtFjhp914IzLrLuDmOW7A/FCp+KTblzXFZLINo4xDasjnmGbRlGG35ZmhPbVcC7XoeAdDCR6bxkR4+cn/WGZuHjw4VUwDpMomdsYlFiHQxCNpBqHrqVHQdNnMRgjA8jV1vMRtdX57r8RC1ZmNTqjUdRWs6PjZm5iMEo7PIGGehftuQ+82p1vdn3eZ81GEG6jC5CJNm2wWPcdZVVd1eFbebsm7wJtCdjyTNgB760MMI9PANoYddCKjaX9vYu2LYoQBHk4j101CLkAxVbCGBipsFVQtemwehKm0eQJ8HUPEgbMFFU5TEV11PJLk8nZjglTW6Rzuu8umVIXBA9ML0QlBk0Av59EIReqE3pBcK0Yt16IV8eiFFr7D101CLUAwB5VLtKZAjtaG26RWq0qYX8umFFL3CFlw0RV+JXhniEexXj+BEE4sJEx2pwyTbEgxOBiZyOK/m680TtizsEwtHiJVWjpxIuSNKItQiGjEKG7R6nwv8kqLHxPXd2sIEbBMN+0TDimjh0UxDLcpOMB8q1eEW9rmFFbfCnV40RV+JWxmCG4wpbiluzd6bwJ2grUL6BIIRn2AkQrCootd4R0VO0CqTgluLVHDoGIat29ryOJZOdpdy/DgCJAFu3jfyd5B5U2dUiieBUh2eEJ8nRPEk7ImLpugr8SRD6IP9Sh+K8gSF1qCnHaN8/Q9GBECYUABfdJ8LiHqhYxT1lx+qlp+w9dNQi5APhRBA+YcgwAlucyxDWYS+tAip4ljYgoum6CtxLEO0hP2q5TM5lneW8rVMGBEzYULNfFGOBfTJ0FmK+RxjimNh66ehFqEcagmYEu1ZtQ+3OZYhkkJfJYVMcSxswUVT9JU4lqG/wqQAa5RwzTSKHo/qX4rFYlIXy7Jh13Q2+vApg1e+CgsjMixM6LAZKrkzCBFo7psRwiR46TRxfTkMGiGWdsJ/7vOKK16FrZ/CkLSL2oLCLFSswyVff4VccSnc64U/jhfnUoayC9PSbhaX9MElSSRf0oURTRcmRN0MIpljVIpFUxiQV0OYZ6iw0JdhoVCYh4d30RR9JcwzBF6YUnj1TT0b7+Yy2EQQi/1/yrourssnHJWlj3tE3YVJebdzEvZQvyJk8NvH918GP/C/gnAH1FpMhx1lJ0PUhb6qC6WCOzyqi6bo68CNMuRi1C8XT2YfLeIu9YY+pvgoSD8NTtWhJIS1UYIgCqe0AE/ORxHBuLEshrjuv2tbMrKeunZ99R531HuUoeUiX8tFYH/WHY0PSYaWi/q13E+b+7JeXxWbEnrZVqIJ72z6DBRoPLnQ18rMXsRIyptbt0iqkZMZqa+Ht9Np0K+rudSmZKm112IWOCL6XsvJwUpJlDoTwGUeuUi4WceINFYxZjIFMBOMmidAfLzrYgpcPQq9drY5Xdt2NSMFlcSF7nTczIf2XNkmOpldS7SzVRwwqJkwu/dNFIA2aU/Rm0ihPjYZ+h5KZodBYDPRADHJhZasQV84QYl5JEW87YiodNcdQIaIhPpFJDPfbKqY5FYHJzoy09egxMow8QGR0IBgZ0C9yV/+gDLUDtSvdpxcZplOg6Z3FoyobtE13QU4CPWY7o7FMdvPLz8OutMOUwrNPsoksXmIRD0H/RuGb1vGgRnZ4xvpUYyfmP7yd7mKHJxcX96qgER3SUhmQSCE7FSXUvCta57imYxjJZJJzzw7TaTHRzKwcgZ8lDxc/rKPcMZZDIOUj553dx/3DwZd/+BOJsMpTt7d/7p/Mg5GGMb8szMid4gJpM/i6MGla03GgQOjDGtQ3JroVt21Jif/GmdYg+PW5O+XOGO/xKSfyZ8vJ7tLsDppmbVZcGqOTNB+Tp7lT3A8WRqns6WnOHHrYCZOKCNplqrZn4Rso4es3vumTMY+j2l62c2ScLCwYrOQGasK9QI8TPfPYrogTtxpqIqHTw0oflsqu0JB/hneXleEe8Pip5ChmLGpKHbWSSk65+9ZU0762Z3qcKtiD9BKBT1/bNWEj8pPqoHe8BFnXDJg1g95/Maze8XQoA4JyICd+bAzBXvkZh0nrhlUxZeEnSVhVwMM4s5aCIXzvppibdhVQCg4a6PeapMp1PuzrnGG7I95P+rdu6XBFlxBMsDlPrhcgRu5p8QJrV9VfElweRpcQYLg8u6khqAbVc+agm14OVCzhHVmNffx5Qrf/hxrnBFZ4ERk0YMvwjQDX+HjKxS+kTtCnIgoVMWXxFck8UV6XgbwFSF8SQBfEcZXSAYhasMrfHiFgrc/qxpnhEc4IzzazbB74uT1E6ixVOBGcjVxIhRSFZ8J7u8D1fGhmh5ak5RNtCCDGvE59jT088QEIhnBFUkGVxEPZ00f4mvaBCgPR9IVSSKYUhWf72EClIdFjocJ8D0c4fBo51fFy7K+Mz97Xw+uq+8rrdPs7bx1P7ZH/N0pMr+vbn8D353a3ys/NmR/w//vor6br9aDRXmrGgVDrg5GtQXQfthU38yvmb9WGwWu/WFzWdyUtS6gvr+tqk3zQXew/ccJjv4PUEsDBBQAAAAIABFNaFxFkUIFbBIAAAZ0AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDYueG1snV1pbxtHEv0rhBZYZLHIqO/DF+CYh3iIFC0q+1mxaVuIJGopOk72128PyZHJqteadgLs0qJev77mdXVXVY9efVutf3/8slxuOn/e3d4/vj75stk8vDg9ffzwZXl3/VitHpb36TefVuu76036cf359PFhvbz+uC10d3uqhHCnd9c39ydvXm2/u1i/ebX6urm9uV9erDuPX+/urtd//bK8XX17fSJPmi/e33z+sqm/OH3z6uH68/Jyubl6uFinn06fWD7e3C3vH29W95318tPrk1/kiytt6wJbxK83y2+PB//uPH5ZfRusbz5OUs2pI+KkU3fut9Xq9/rXw4/1V3Vl98vOn5cPtzepenXS+ev7Pzerh8ny0+bd8vb29ck7fdK5/rC5+WN5kUq8Pvlttdms7rbNTp3YXG/Sd5/Wq/8t77ctWt4uEzg19WGLTlQ7KPjdjqmuKf/bfT27Buwa9LYevP9uR+LtdthOn7p++O9mSPrbGUsz8Nv14/Ld6vY/Nx83X16fhJPOx+Wn66+3m/erb2fL/SxsR/XD6vZx+/+dbztsGoEPXx9Ta/ZlUwPubu53n9d/7ifvAK9UpoDaF1CkgMwV0PsCurSA2RcwtECuD3ZfwJbW4PYFXGkBvy/gSwuEfYFQWiDuC8TSAlI0MyeKizxNNp3tfJFmumXxfMtmwmXxjMtmyiWbc5Er0ky6ZLOeLdJMu2Tzni3STLxkM58VVDP1ks29zBVpJl+y2c/VoprZV2z2c7WoZvYVnf1sJU9a307+6W5R2a5I3evN9ZtX69W3znqLr1ce5Rqep7UordEfasQv9XqX5JOGM319c18blMvNOv36JjFu3vzzH8pLJ1+mz6ClSJ9SWWm2nyb43c9K17+Xxqj6Mwpd451VIWy/t6L+1FKLuP10Ou7Ke7krb8SeL2w/ddx/CgF4c/XXePny18lsNr662NFsmy203X7GGNW+WfHlq9NNGra6i6cf0v/ScD2NmdqNWf205MZMbcesns/MmOXHyEpPKt9SvttTqucold1RuppaOWPCy5/u9L8AXbeVThkl1I4m2hxNr51GhW2rlJDSt9H1W+l0sNu5T8Mndp30VufoBgVjZuJu+E3UT9OQoTv78SlQXslEpxDdsIVO9r9Pgg7emieJZAhHBZNhmjndKa6EdlxAW3ezpk3SKaWdtNCqH+3+eXs7RdguJD9EOy15iILbPzxyv27pHN3shyepfpj2y6FVe3qDloiLkhF1LmynKsnS/Qj5vIW8d/V+thuKhk68/GnbCaleooF430LYH0upfkq0qPBlS+H36VyQioOSi/ZqtclVe1VSrTbP2A7dbjv0to4osg9INL6eqsOPztmwfoxjWlo7EPBMk0x7k8yuSTlzdjb8WWpks3blauZUrj6Z/vFGxkrLV6d/HNoiBKtcOEb1IMqrY1QfoERl7TFqALn0MegMgUKoLGn+EOCcqpw/ho0QndKVMMe4McBpU0lCNwEwEytLRu0c1SpVpciITBHOMdhs/ySooxkloIsS0ByBjiHvdxClDyBWOCWOYZd7Jn005VJEMp8LRGfSFvIYdpWhk/EJdyQf2y4fu2M0z8nnEulnX9AeNEWlZ94R/QCYrCIZzR4iqwTh6gOUqDR5qgaIi6gHQUQSP5mVIcCFWv3i+3+kJyPUX+MqS/oyBjibppIICaCMrSKR5TmqVPnKkX5PEc5WIRIh7WHuWSGVgOYIRIS0h/jDgQjJQhMh7WHHa2eyZ6TtC0Dn6jMaERKmUwcjdiQk1y4k12qHtEA6cmjpZVPcBTBReTLePURWKTJGfchlqI4gFwGdIZAMVaB2COCSRaB1jgAspuWCwMaILVVKzRDqJ1FOWbumqJvcdM9ciQUCINKqOeIhsnHcYOh0dKCycdBgaGrYF4jOWUdlg+mUzsjGt8vGt9ofbZFsPFjKHHvuughWUdVAUCDGoA9Q6REgAzmAXESAZwhkNVuvhwBnDVcNotOiiqTaMcBpz5b/CYAlCyZI485RrULQwZ0imGfr28yX2JwS0ByBiHg8NxLGeWokLj00Eo7uYxeILkYmHkxno8PiCe3iCe02xyHxBLCcpW0FeeK7CFZJMnM9iAoE1YcoUuEAgFRFmM4Qk05HHyoegPOWrRAjSCeZsMcA59KWkcAmAEZtTlnDpgAWKxXi4X9ERKHE/gAQtT+Ih0gocINhNDuwXAZsf4ilXiA2pyVVUIbtYEE8UlBsV1BsNz8eKSiCVS3NDt21IRg7sfQgip4Y+gAlqkBQA8hFQGcI5F1l6PkH4KytFKEbITop2XFqDHDpNGMJbAJgOpl26j0AsKjY8jRFjdNspmaxxP6UgOYIRMQTkcHg9idCg2HoNmEB6KzWgaoH0+mYcR7U0cs2+dSYNgsUkX72BY+XXcNW3S7CpZ0NWT96kI5LCLMZMp4DBKMuBFij5ge3IQI6U2mydo8goeRexDEkrO1H3jBMYMeJngobOoU4thDNmkfjeWOEUNQaQSaiqD3m+ADjjaeSariOTYiImjRsgQiNkMyRkCEMMXMkkrJAVbLVKhkJVSXRSsy2Rl2EE5VzVFWIjm7a+hky6k5AMCYqVKHQlfVUVABoNNPKCOGSkaCP4RjymUpRfxzsK9VRWdumsLOBGeFZg3veNCEUExJiokKS3Jho5biQJD4c0W4uIKEJ1FN+lSMUmeORLEh1kKrVPNV7RyAkBRbjyOIhXYijKgIY4NxGMEEPPwNYIT1zn0GUC8xhNURAx90oI0hofOWoiwF2lm+MJ7C3VEhljZsinNaV8VRIqsgiARQTEsAoKiTFDYiVgXoFLhsucqihfm1IZ21kMsJ0websUUHUV+p2e/QWykgjR1flLJURwKWzDfXTQTrmNetjNr7Lg2yeSgmhTD2iVEoAaH3aCVApIcI0KNSrNIaEktmHCewulVJZ46awcUlK9LTU4FpsEkAxKSEmKiUNzktCaCYljU849FSxgIQq0qDTVY5Q2IyYCvIVZHvCgnn778msv4CKArFwpSpP/d4IJ9nS2SuD9RFMMBkPMjCqKNSFtMmTNGaEgOmgo0T+oDOCHVI892AMga4SNIaEcNay9fkcV8wM+BTiDOObyaJshiLUHM+ypSpDOQjesUisxEkIjvpxFojQyci8ehlCexDPO1ZZQVqDbM9rsNA13hQ8OjGwx6eLYDwg2MMwGs/sYxiN2w4QTNKcgTOECrXnmMoL4LSrJDtDAZxDikJ8ulJMUaizVE1lbZvi2aLpTbMG1mKuAIqZq4KMhgZzaF2Ujcy91+CodaH+4QUk9I45+DKE6kCZx0IqSGuQ7XkNFnrIJYqZs/MTTDJgISYMo2kNGEXP8AMEM9xKITKtqkDDTAjoHcuJG2FCy/ZgYwhUhqc3ICATUlnrpgin+GZz1jwOLWapIMcBMlElgbQEowXLsmu4qFeP+SIQn2VZe1cZvuCigKlfx6oqyHqQ7WkP4T1UFYy6M093F+HSVowJC6Yr0NN6H7OxjKEMG1UWQEVb+We2dUNUxqRtLztcAVzg2QZjyBfYUXICO04lVta2KcJ5ttbNZFEmBEIxhRVkQkiQu6CC414KnLyg6H55gQi1cCyTNUMoD85qx6oqSIeQ+/jwj6WnT2Z/Oz1dFkSYZUuIeTL7WWFnPgpw8hAzggmWh9tDMMmMTx+z0bTyAYbRNCeECpIZnyHCKR6OHiFcMmaZRXmvdMRt2YI5gThTeZpqi3CW92kKcZWhGYOyKO5chJojFPNKglixqgeQ6h0Gi5Pc2SEPECa903jBFSaUMWRyN1RB9Fm1RJ+TtDS8+6FQrDWdG2gCFMRVmmoLwzzdnyIYyCDEbPSUB1GBb3eHCJgeWHpaH0FCFVigYoyA7JCHQJpn2J8jXORpLVPYOs1mbKZArFhRX+0FQnE9QRTRkwJhYiu4/VQwTCytoHFnSGg0c01mCE0uFVcVxJ1VS9y51hNMxlUolOkrR72SEEePXj2MotPTxzBJnZIIZtimFJJZy+6DDRHQarYmjCBhrQCqJhjyNszVMcEVUzlBNseaN4VAEHxWRcHnItQcoqieQLTYas+ckApHiyMNSC4goZcsbpYjVDk9ldy0bwk/13qC4WcFw600NtWFsMowPeGYsad6ggFoemQcIJji5gnVqS1bEoYIaCILr40goZQse32MgE7wVA5YsWdelnOEi76iC/cUjwvdOc9UURC6CDWHKCooEDc20lIn/KXCcWPDAtGQ0ABBYUJtMm4TVRCIVi2B6CQoM4GCglHciu5muxDHdu09DGP3fREs7W2keObAMkCFmLWCYVrFMriGCGhCRfOvR5BQKZY5NkZAq5isJ4UVnyNcdGli8s6gKSoTeEZvAzs6DvGtYEkkew5RVGkoqAyuZDU4YmkCN12I0PEodYbQx4wrRRVEqZX5O66UYe9vu1JUQVBPtQT1hr2fpX0H9Q/iRIrvKbsIJ1mgu4dh9DZmH8HAhWUE0yynC3ZBKOahGCKgszyuBztRxxlociQEpi0qfWAnCGjSKstEDwkjs+VTCNQVXTNnDa5ll1qCmiMU9aI0mKNNpWI3vi4b3LEyhZaOSh0ROsMifBlClXsBgCqI8KmWCF9SFXZQKhSo8uz56UIcPQj1MIpuZvsIJlj+ywDBFAvywTodF/wQAY1jt79GkFBYuj0eI5wNbKMxQbh0iIyOKgrggqNjPEUwzzIDZs0T0bJHBShmX+cQReWEwnJBsTBfw0WdKDSGskCECcXlhAmNzsmpILSnWkJ7SU4avfznnULRs8juunQhjsupBNVHKMF2cwME0yxVElaZnkLqQRwioNGVYhYKRQrVs5vDMWyE4KvSJNcIes0M4VKn2PEPXsJm7qiZKgrvFaHmCKWptkA8zmrFHSowHicizXdYQEIXWQw9R6gy+fyqIMCnQqupMl1sqwKwB/wiWRfhuLO8l4GxAyCACeZ5GWA25lJBqBDAoQ8ArWRP9ggSisCAYwTUgp2eJ7BiHos8xxUrdr1iioFsfZo1z0WLwSpBzRGKGawA7IsO7JJMw0VTJakzagEJPb8kkyG0NpMqqQoC1KolQJ1EhVMlFQopexoF7kIYt1f44jIzWBBGdx0DBNNcUYjM8XD3EAFrJyW9JgMJpWOZMmMETGc06lScwH54nn2McGj/h9qHvP5FUWmEAhvAgvvQCgSRTWSvn7lUOIjs6Y30BSK06djF9IQJXczoSRdEpXVLVLrWE7wTrXFUmt5j6kIcy1zvZWA0twvBBEuhHSCYZq8lhHV6yzKeh5DOslujI4TzLHw3hjBPYZNMrfT0eA7pJL8KjXCWvftkpkvCzRdFqDlEETlpdHc5CCYnnYkh05jPAhFa6dhl6Ayhjhmfvy4ISuuWoHSSk4PnKY1DnNT2diGOeap7GEY3X30MoxoeIBgL8J1BMmf4lg8BnWLb2xEklPzlOWMIVIH58iYImHB0i3aOcFGyWwRTWDFPfp7posB0EWqOUIZqCgemWVK/hnFkEamraAEJvWA+ihyhyASmdUFgWrcEpmtNwcQpjYKcII6GcJJdwexhGNcUjExTJ+oANo5leuAuCH79DAGT9uh7aEawE0lTnkoK4Gxkb9OZIFw6bdG94TnCRVNFnQ8uTmHnJVP1TIPIMo+dIRRQV0GUWoOgslOGW6xMUJm+C3gBCb2mV0avMoQqd6DSJS/JbolS1+oyUF0o9OvYfqKLcOzmcg+jqEnoIxh4rxRsG9//IZQJ7IW/QwT0gjnAR7AThmdNjiGhrtQzvsIJKmMUe/PjOWyE5L7CKew+vyo3a3AtxqskgD2HXFReIJLsPHewNzia5U8fmgUi9EqzrN8cock4AXVBaFq3XKCu5QWzFDV833glmfGC15iZvOClWrri9jEZfc4HmI3ZLoQSjjo2hghXLyQ05xfhvHz2VVJjVKZ2BgYqLnjNm11qPUc44/lVT9x3ZhZnuuRK9EURao5Q1BeowS1nLQXzBWp8y1mxdyFCQg02hphQmsxtT12QYaHbMywc9AU2BUmCBXV8dRGOJ+32IB0QF0ywoMvyALPRYDBsWvTsAvMQAa1m68gIEtZvgmI7QwD0gl0Nm8B+6MoySSE+x99ABRsoWIdnDe75TCqEApIqSK9oMK1nLZgNIQWLB0NC7/huEKdXRHaB+vTgr5XdLdeft39/8bHzYfX1flP/1ZKDb/d/mFK/uNr+3Tr6vQwvrmSAv0m/QN8r8+JKbf+g3en3qnd/IfP8ev355v6xc7v8lJpRvx/+pLPeaXz3w2b1UMu9s/sLktt/fllef1yua0D6/afVatP8UFfw9Kc/3/wfUEsDBBQAAAAIABFNaFx9xefWogcAAONgAAANAAAAeGwvc3R5bGVzLnhtbN1de2+jRhD/KhanSj2pPbO8DG0cKSFYqtRWVe/+qNSrKhywg8TDxeSU9NOXBQyLMxNhPOHWxTrFMDszv3nssLvgvat98RyHHx/CsJg9JXG6XyoPRbH7YT7f3z+Eib//kO3CtKRssjzxi/I03873uzz0gz1nSuK5pqrWPPGjVLm+Sh+TVVLsZ/fZY1oslUV7aVb/+SlYKswylFktzs2CcKk8l8c8SeZBoMzB9ma/vfpBLQ+krfWi7TdIy0W/5bvv3r2rJCPNbQgE1tgBZP/9/sc/fw+Dvz5/W51+fg/zLlQEV3U0Ir7vX+WS5o3vr682WdqFQLOV+kqpy0/C2Rc/XiquH0frPOJsGz+J4uf6ssYv3Gdxls+KMvilcsav7P+tyaw+43nRyEmiNMsr5bWGYz03eeTHnL5uJHQK8u26dKK6qo6eFmOIwAgTqJeH44gCHWqA6hCBPSFdlM4U4rqnC0FdxUzLdM9FZdwwwx6BCnX4WF+9KpDUb1Z10KWY6x5bPEgeis+pDlGefRY+dmt6GmlE7ozFDSMU6FkmG+HBE+SdmTK3K91SXTqAgAc1aov1swTatuoZK0KBQKEZdKc4pbIP8uGb1gIAFCMup4MA0ljDVrptGWg/qv7wAUsUx+2ARWdKfeX6aucXRZinq/KkYqouviDNmu+fnnfliGWb+89MM5XBDPssjgKucuvC0NcYYS7IPFOb5620OuxrjECobWXyD6CtIxBq07yFeWsB2joCpW1tBzq2rSXQarvzFrC2hkCp7YZ/IG0tgVBbNzw87gEtgVCb6Vq3oLaOQJuTd7oG52RDINRmmOYNWEs6AmXc2uHacdxaAmXl0ssCdQdVrpZAqc32blZQLekIlP1NXRkrG+pvLYHUtlKsB9p2IFDapnvIPaAlkN7f+AHe3w4E2qqse1AP6Ai0cTM9B45bQ6C1zfYYbFtDIB0pLFwV9GRLoPRkOwk59mRLIPWkp95C94COQKjtziqnQ5BtHYHSNtczPBeyrSVQxq0c6sBVuSVQ2ubwD2RbS6AdK9vweLIlgNqqP+XcZp3lQZi3sxu+Gltfur6Kw01RsufR9oH/LbId15IVRZaUX4LI32apX818Dhwi56xauV8qxUO18t5fbqiOChtv2ugYyFG1reAMZChbHnAP5KgbjzDMrY5TDBM4hhkmMAw0TOAYbdhddZximMAxzDCBYaBhAserhvWT+PSsGZDt8+ly+AQ0Yw0+sZtI5BwC5M2Xsjbeh3H8kcv7Y9MWSL7687QRHohVj8PS9mtZVZuvtZj6hMsXpdWyBbGOM0rubBd9yYrbx9KatDr/5zErwt/ycBM9VedPmxYAJp110jVRennd3+3i55s42qZJWBs/WOH1lX/gmz1kefRvqY0vsN2XF8JcmX0J8yK6F65wFz1tcJhaB1MXYWpvApOn7QiQQqQMEaROEim9k26+faTOd0EvWS0SF2DSFyTSjU66JZeDDadDaXYoF2+PsqrGI/JA8KUtly+bFzwuyZn6ZdToaQsUBUznMmAy9UJwsgvBeSG9iOmy4TS+1t3ypNIpoLQEbxqSwdRBZ7IJaudYb0p8u4Qzk8mWmnDMZUOJOFO2mAtVcyHAnGDAObq6i+6cYPQxchZnT1uOxk6FHHnLUfNa+Fdx53igTJUYqRh5xiQGinR3yToS3N21CaYap83WTdCbU+AcG3SmyetPePwhszfF5Jxgfnn+uucLlGOXJoVYMX1aLxDUO4mDxcSyPMGs8PyceoGSIqcm9gLFPVTeYDFhbqxNsL5EUKeOUY7NqeZnjY0frGn9QFGp5A2XWPYnWAwkSKpjlLQPPqdwAsXKrTZBoToNJzLwk7iH9iqJdE9AerPRxbRIKYbS8s5G1VdQUgx7rGm9QLA8PMGC5liUEj8RQBJfslc7kMQ/RkmQ+BM7YWysVDyjaN8gonIxPKicoDeMXPdhMidCby4h8RMypNse3/wp7lfO5Q0uJEt+JFhU5UUIlvaafEmDJdn7eUiwjlFSBItN64WxwdLE5wmyzYNEnOIEWN6H25owZp3iodzoN6w0YabSK/xv89L7iffpxWUA1S8DZrODm/xA1cuA2XtuLPEktTfelRinGHZx/VC69dhmd8MLyNDFhXQlEaa40CnZQoq4ICsMbyVGKfH0FnvtRjacSND1CebJJ0UddOYUKEd7U2ac8Cq5dDCxTiQD0HnzC2ThZ869Hzm3V2d8R72l8ivf8zfu9M7Wj1FcRGlz9hAFQZg28+3ut86l+MJfx2Ffftk+CDf+Y1x8aolLpfv+SxhEj4nTtvqN+6Jp1X3/mVvIqu07qt95l7qiNAifwsBtTvPturcnRrfZ4DGl233tJQXjqWkwhdMwPRgCjKfmwvT8n+yxUXtqGobNBik2ymOjPDUXRHGrD6YH5uH73MKWOo6u1/tyQh6td+t4gcDF/GZZ/B8sDcPGOTA9XNNpvsajjWfI63mAxfS1DMEsxTMRsxT3NafAfuMc9b7GL6ON6eEcWBSw3OH6YT08p2AeXT/sAQNhw3owTnEcjMJzEc5Ry0K8Y/EPHB+sl/BN1GEKp8EI+NbrMIX3RpyCIThs5A5R9Hoz4aP70fxwn5p3/63B9X9QSwMEFAAAAAgAEU1oXJeKuxzAAAAAEwIAAAsAAABfcmVscy8ucmVsc52SuW7DMAxAf8XQnjAH0CGIM2XxFgT5AVaiD9gSBYpFnb+v2qVxkAsZeT08EtweaUDtOKS2i6kY/RBSaVrVuAFItiWPac6RQq7ULB41h9JARNtjQ7BaLD5ALhlmt71kFqdzpFeIXNedpT3bL09Bb4CvOkxxQmlISzMO8M3SfzL38ww1ReVKI5VbGnjT5f524EnRoSJYFppFydOiHaV/Hcf2kNPpr2MitHpb6PlxaFQKjtxjJYxxYrT+NYLJD+x+AFBLAwQUAAAACAARTWhcAj5mZyECAABRBwAADwAAAHhsL3dvcmtib29rLnhtbL2VW2sTQRTHv8o6FPrW3WwuYMjmxaoNFBts6WuZ7M42h84lzEya2seE4gWKb14QRAVRKfjikxr0w6xpHvsVnN11cdPVIJLmaeacM5z/75yds9MYCHnQEeLAOmKUKw91te7VbVv5XcKwWhM9wk0kFJJhbUy5b6ueJDhQXUI0o7brODWbYeCo2chytaWdN4QmvgbBjTN27AIZqN/x2LQOQUEHKOh7Hkr2lCCLAQcGxyTwkIMs1RWDDSHhWHCN6bYvBaUeKqWBXSI1+AX3dgy5gzsq8WjcuYsNiIdqjkkYglQ6OZHkx4bxkJjDqdXX4hZQTeQ61uS2FP0e8P04janCzpWR9CFb0ybW5b+0UYQh+GRd+H1GuE77KAmNAbnqQk8hi2NGPDT9cDr9+HRy8nby6MXF+MFGa/rp68X4YVyhkWwFabXaYOZ6J+tgArIVJMBLg9vcKsC5c+DcpcK1bhbgynPgykuFi4Zn0fBzNPwejV7NMlbmMFaulnFy+uzHt9fR6GWGNs5xVedwVa+W6/z98/MvT6LRm2h0PxqdTR4Ppyfvcmi1OWi1ZIKzsQ1ICJwEd0zaWeuX0t4R5WytLYHrvR3QlBgGKuI/TabloObqX0Z09dpKqb5Sbti5vP8nUiqKZKO2OBG3KJKNzOJEykWRS1d/cVpVo/XHq5JKuJck7NnLYDL5bWnFS/IJ3Eq1dN08G31KbxjfFt8UOMhehOw1a/4EUEsDBBQAAAAIABFNaFxlf1pqzAAAAM0EAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHPF1L0OgjAQB/BXIX0ATxDRGGBycTW8QIPHR6C06Z0R3l7UAZs4uJBOzV3T//2Wa3rFXnKrB2paQ8Go+oEy0TCbEwCVDSpJG21wmG8qbZXkubQ1GFl2skaIttsE7HeGyNPvzKCYDP6TqKuqLfGsy7vCgX8Ew0PbjhpEFkEhbY2cCRj7pU3wPsLNnCyCyy0T9nILBfgGRQ4o8g/aOaCdf1DsgGL/oL0D2vsHJQ4oWRFEPPVIi+ZTO+MPK47n+S0u09/lp+nu9fFlAOf3yp9QSwMEFAAAAAgAEU1oXHD5tdwuAQAAdwYAABMAAABbQ29udGVudF9UeXBlc10ueG1szZXPTsMwDMZfpep1ajMG7IDWXYAr7MALhNZdo+afYm90b4/bbpNAo2IqEr00amx/vy+2lKzeDh4waoy2mMUVkX8QAvMKjMTUebAcKV0wkvg3bIWXeS23IBbz+VLkzhJYSqjViNerJyjlTlP03PA2KmezOIDGOHrsE1tWFkvvtcolcVzsbfGNkhwJKVd2OVgpjzNOiMVFQhv5GXCse91DCKqAaCMDvUjDWaLRAumgAdNhiQseXVmqHAqX7wyXpOgDyAIrADI67UVnw2TiDkP/vRnN72SGgJy5Cc4jTyzA9bjTSNrqxLMQBFLDRzwTWXr0+aCddgHFL9nc3g8X6m4eKLplfI+/zvisf6WPxUR83E7Ex91EfNxPxMfyH328O1f/9VXYrqmRyp74ontv1p9QSwECFAMUAAAACAARTWhcRsdNSJUAAADNAAAAEAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnhtbFBLAQIUAxQAAAAIABFNaFy5FZzL8gAAACsCAAARAAAAAAAAAAAAAACAAcMAAABkb2NQcm9wcy9jb3JlLnhtbFBLAQIUAxQAAAAIABFNaFyZXJwjEAYAAJwnAAATAAAAAAAAAAAAAACAAeQBAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgAEU1oXH3SYYROJAAAQe0AABgAAAAAAAAAAAAAAICBJQgAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQIUAxQAAAAIABFNaFxR283yMSQAAAXtAAAYAAAAAAAAAAAAAACAgaksAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxQSwECFAMUAAAACAARTWhcjV3qdUUkAAAs7QAAGAAAAAAAAAAAAAAAgIEQUQAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1sUEsBAhQDFAAAAAgAEU1oXAw3XgeMIgAAl98AABgAAAAAAAAAAAAAAICBi3UAAHhsL3dvcmtzaGVldHMvc2hlZXQ0LnhtbFBLAQIUAxQAAAAIABFNaFzfzAVXpQsAAKhAAAAYAAAAAAAAAAAAAACAgU2YAAB4bC93b3Jrc2hlZXRzL3NoZWV0NS54bWxQSwECFAMUAAAACAARTWhcRZFCBWwSAAAGdAAAGAAAAAAAAAAAAAAAgIEopAAAeGwvd29ya3NoZWV0cy9zaGVldDYueG1sUEsBAhQDFAAAAAgAEU1oXH3F59aiBwAA42AAAA0AAAAAAAAAAAAAAIAByrYAAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACAARTWhcl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAGXvgAAX3JlbHMvLnJlbHNQSwECFAMUAAAACAARTWhcAj5mZyECAABRBwAADwAAAAAAAAAAAAAAgAGAvwAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAAAAgAEU1oXGV/WmrMAAAAzQQAABoAAAAAAAAAAAAAAIABzsEAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQDFAAAAAgAEU1oXHD5tdwuAQAAdwYAABMAAAAAAAAAAAAAAIAB0sIAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAA4ADgCcAwAAMcQAAAAA'
    )

# カスタムシートの行マッピング（F列に数量を書き込む行番号）
KIT_ROW_MAP = {
    'log_wall_m2':    10,   # ログ壁（LogWall）[m2]
    'floor1f_m2':     11,   # 床板1F（FloorBoard 1F）[m2]
    'floor2f_m2':     12,   # 床板2F（FloorBoard 2F）[m2]
    'ceil1f_m2':      13,   # 天井板1F（CeilingPanel 1F）[m2]
    'ceil2f_m2':      14,   # 天井板2F（CeilingPanel 2F）[m2]
    'beam_m3':        15,   # 集成梁（LaminatedBeam）[m3]
    'panel_m2':       16,   # パネル（LogPanel）[m2]
    'terrace_m2':     21,   # テラス（Terrace）[m2]
    'euro_door_cnt':  22,   # FINドア（EURO Door）[本]
}
KIT_VMW_START, KIT_VMW_END       = 44, 79
KIT_NAGAI_START, KIT_NAGAI_END   = 84, 102
KIT_VELUX_START, KIT_VELUX_END   = 107, 115
KIT_COL_F = 6  # F列

# モデルタイプ→シート名マッピング
_MODEL_SHEET_MAP = {
    'HI': '積算入力（HI系）',
    'IE': '積算入力（IE系）',
    'LO': '積算入力（LO系）',
}
_DEFAULT_SHEET = '積算入力（カスタム）'

# VMSD-1 → 1810x2019 のような特殊マッピング
_VMSD_MAP = {'VMSD-1': '1810X2019', 'VMSD-2': '2419X2019'}


def _detect_model_type(model_name):
    """モデル名からタイプ(HI/IE/LO)を検出。不明ならNone"""
    mn = model_name.upper()
    for prefix in ['HI', 'IE', 'LO']:
        if mn.startswith(prefix):
            return prefix
    return None


def _normalize_mk(text):
    """マーカー正規化"""
    return re.sub(r'\s+', '', str(text)).upper().strip() if text else ''


def _extract_base_code(label):
    """Excel品番ラベルから品番ベースコード抽出（サイズ情報除去）"""
    s = str(label).strip()
    s = re.sub(r'\s*\(\d+[xX×]\d+\)\s*', ' ', s)
    s = re.sub(r'\s+\d+幅$', '', s)
    return _normalize_mk(s)


def _build_item_map(ws, r_start, r_end):
    """品番→行番号マッピング: (exact_map, base_map)"""
    exact, base = {}, defaultdict(list)
    for r in range(r_start, r_end + 1):
        label = ws.cell(r, 2).value
        if label:
            ls = str(label).strip()
            exact[_normalize_mk(ls)] = r
            base[_extract_base_code(ls)].append(r)
    return exact, base


def _match_row(marker, exact, base):
    """マーカー→行番号マッチング"""
    mk = _normalize_mk(marker)
    if not mk:
        return None
    # 完全一致
    if mk in exact:
        return exact[mk]
    # VMSD特殊
    if mk in _VMSD_MAP:
        sz = _VMSD_MAP[mk]
        for k, r in exact.items():
            if 'VMSD' in k and sz in k:
                return r
    # ベースコード一致
    if mk in base:
        rows = base[mk]
        if len(rows) == 1:
            return rows[0]
        for r in rows:
            for k, v in exact.items():
                if v == r and '(LF-3)' not in k.upper() and '幅' not in k:
                    return r
        return rows[0]
    # 先頭一致
    for k, r in exact.items():
        if k.startswith(mk + '(') or k.startswith(mk + ' '):
            return r
    # 部分一致
    for k, r in exact.items():
        if mk and len(mk) >= 4 and mk == k[:len(mk)]:
            return r
    return None


def _collect_kit_quantities(all_elements, marker_stats):
    """all_elements からキット積算用の数量を収集"""
    q = {
        'log_wall_m2': 0.0, 'floor1f_m2': 0.0, 'floor2f_m2': 0.0,
        'ceil1f_m2': 0.0, 'ceil2f_m2': 0.0, 'beam_m3': 0.0,
        'panel_m2': 0.0, 'terrace_m2': 0.0, 'euro_door_cnt': 0,
    }

    # ログ壁面積
    for e in all_elements:
        if e["部材分類"] == "ログ壁":
            a = e.get("面積(m²)")
            if isinstance(a, (int, float)):
                q['log_wall_m2'] += a

    # 間仕切壁 パネル面積
    for e in all_elements:
        if e["部材分類"] == "間仕切壁":
            a = e.get("パネル面積(m²)")
            if isinstance(a, (int, float)):
                q['panel_m2'] += a

    # 床（1F/2F/テラス/バルコニー）
    has_2f = False
    for e in all_elements:
        cat = e["部材分類"]
        a = e.get("面積(m²)")
        if not isinstance(a, (int, float)):
            continue
        if cat == "1F床":
            q['floor1f_m2'] += a
        elif cat == "2F床":
            q['floor2f_m2'] += a
            has_2f = True
        elif cat == "テラス":
            q['terrace_m2'] += a

    # 梁体積
    for e in all_elements:
        if e["部材分類"] == "梁":
            v = e.get("体積(m³)")
            if isinstance(v, (int, float)):
                q['beam_m3'] += v

    # 屋根本体面積 → 天井板推定
    roof_body_area = 0.0
    for e in all_elements:
        if e["部材分類"] == "屋根":
            tn = e.get("型式名") or ""
            if "破風" not in tn and "鼻隠" not in tn:
                a = e.get("面積(m²)")
                if isinstance(a, (int, float)):
                    roof_body_area += a

    if has_2f:
        q['ceil1f_m2'] = q['floor2f_m2']
        q['ceil2f_m2'] = roof_body_area
    else:
        q['ceil1f_m2'] = roof_body_area
        q['ceil2f_m2'] = 0.0

    # ドア・窓マーカー別カウント
    door_window_counts = {}
    for mname, stats in marker_stats.items():
        cat = stats.get("category", "")
        if cat in ("ドア", "窓"):
            door_window_counts[re.sub(r'\s+', '', mname)] = stats["count"]
            if mname.upper().startswith("EURO"):
                q['euro_door_cnt'] += stats["count"]

    q['door_window_counts'] = door_window_counts

    # 丸め
    for k in ['log_wall_m2', 'floor1f_m2', 'floor2f_m2', 'ceil1f_m2', 'ceil2f_m2',
              'beam_m3', 'panel_m2', 'terrace_m2']:
        q[k] = round(q[k], 2)

    return q


def generate_kit_estimate(all_elements, marker_stats, summary_rows):
    """キット積算・価格計算 Excel を自動生成（内蔵テンプレート使用）。成功時はパスを返す。"""
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_PATH))
    model_name = os.path.splitext(os.path.basename(OUTPUT_PATH))[0].replace("部材一覧_", "").replace("buzai_", "")
    kit_out = KIT_OUTPUT_PATH or os.path.join(out_dir, f"キット積算_{model_name}.xlsx")

    # モデルタイプ検出 → シート名決定
    model_type = _detect_model_type(model_name)
    ws_name = _MODEL_SHEET_MAP.get(model_type, _DEFAULT_SHEET)

    print(f"\n=== キット積算 Excel 生成 ===")
    print(f"  モデル: {model_name}  タイプ: {model_type or 'カスタム'}  シート: {ws_name}")

    # 数量収集
    q = _collect_kit_quantities(all_elements, marker_stats)

    # 内蔵テンプレート読み込み
    tpl_bytes = base64.b64decode(_load_kit_template_b64())
    wb = load_workbook(io.BytesIO(tpl_bytes))

    if ws_name not in wb.sheetnames:
        print(f"  ERROR: シート '{ws_name}' が見つかりません")
        return None
    ws = wb[ws_name]

    # スタイル
    inp_fill = PatternFill("solid", start_color="FFFDE7", end_color="FFFDE7")
    inp_font = Font(name='Arial', size=10, bold=True, color="1565C0")
    inp_align = Alignment(horizontal='right', vertical='center')

    def write_val(row, col, value):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.fill = inp_fill
        c.font = inp_font
        c.alignment = inp_align

    # 材料数量書き込み
    for key, row in KIT_ROW_MAP.items():
        val = q.get(key, 0)
        if val and val > 0:
            write_val(row, KIT_COL_F, val)
            label = ws.cell(row, 2).value or key
            print(f"  {label[:30]:30s}: {val}")

    # VMW / Nagai / VELUX 書き込み
    dwc = q.get('door_window_counts', {})
    if dwc:
        vmw_ex, vmw_bs = _build_item_map(ws, KIT_VMW_START, KIT_VMW_END)
        nag_ex, nag_bs = _build_item_map(ws, KIT_NAGAI_START, KIT_NAGAI_END)
        vlx_ex, vlx_bs = _build_item_map(ws, KIT_VELUX_START, KIT_VELUX_END)

        unmatched = []
        for marker, count in sorted(dwc.items()):
            mu = marker.upper()
            if mu.startswith('EURO'):
                continue

            matched = False
            if any(k in mu for k in ['VMW', 'VMSD', 'MTF']):
                r = _match_row(marker, vmw_ex, vmw_bs)
                if r:
                    write_val(r, KIT_COL_F, count)
                    print(f"  VMW  {ws.cell(r,2).value or marker:30s}: {count}")
                    matched = True
            elif any(k in mu for k in ['NV', 'NVS', 'NW', 'NHW']):
                r = _match_row(marker, nag_ex, nag_bs)
                if r:
                    write_val(r, KIT_COL_F, count)
                    print(f"  Nagai {ws.cell(r,2).value or marker:30s}: {count}")
                    matched = True
            elif 'VELUX' in mu:
                r = _match_row(marker, vlx_ex, vlx_bs)
                if r:
                    write_val(r, KIT_COL_F, count)
                    print(f"  VELUX {ws.cell(r,2).value or marker:30s}: {count}")
                    matched = True

            if not matched:
                unmatched.append(f"{marker} x{count}")

        if unmatched:
            print(f"  ⚠ マッチなし: {', '.join(unmatched)}")

    # プロジェクト名
    ws.cell(5, 3).value = model_name

    wb.save(kit_out)
    print(f"  ✓ 保存: {kit_out}")
    return kit_out


if __name__ == "__main__":
    main()
