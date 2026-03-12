#!/usr/bin/env python3
"""
add_sheets_to_excel.py – 廻り縁・巾木・額縁の集計JSONを部材一覧Excelに追加

Usage:
  python add_sheets_to_excel.py <buzai.xlsx> [mawari_summary.json] [habaki_summary.json] [gakubuchi_summary.json]

各JSONファイルはオプション。存在するもののみシートを追加する。
"""

import sys
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict


# 共通スタイル（extract_ifc.pyと統一）
HFONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HFILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CFONT = Font(name="Arial", size=10)
BORDER = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
ALT_FILL = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
NUM_RIGHT = Alignment(horizontal="right")
BFONT = Font(name="Arial", bold=True, size=10)

# セクションヘッダー用（グリーン系）
SEC_FILL = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
SEC_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")

# 小計行用
SUB_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def write_header(ws, row, headers):
    """ヘッダー行を書き込む"""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HFONT
        c.fill = HFILL
        c.alignment = HALIGN
        c.border = BORDER


def write_row(ws, row, values, is_alt=False, is_bold=False, is_sub=False):
    """データ行を書き込む"""
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = BFONT if is_bold else CFONT
        c.border = BORDER
        if is_sub:
            c.fill = SUB_FILL
        elif is_alt:
            c.fill = ALT_FILL
        if isinstance(val, (int, float)):
            c.alignment = NUM_RIGHT
            if isinstance(val, float):
                c.number_format = '0.00'


def add_mawari_sheet(wb, data):
    """廻り縁集計シートを追加"""
    ws = wb.create_sheet("廻り縁集計")
    headers = ["階", "部材種別", "壁種別", "要素タイプ", "勾配面", "梁名", "長さ(m)"]
    write_header(ws, 1, headers)

    row = 2
    # --- 種別集計セクション ---
    ws.cell(row=row, column=1, value="【部材種別集計】").font = BFONT
    row += 1
    for mt in ["廻り縁１", "廻り縁２", "廻り縁３"]:
        val = data.get("type_totals", {}).get(mt, 0)
        if val > 0:
            write_row(ws, row, ["", mt, "", "", "", "", val])
            row += 1
    write_row(ws, row, ["", "合計", "", "", "", "", data.get("grand_total", 0)],
              is_bold=True, is_sub=True)
    row += 2

    # --- 階別集計 ---
    ws.cell(row=row, column=1, value="【階別集計】").font = BFONT
    row += 1
    for fn in ["1F", "2F"]:
        val = data.get("floor_totals", {}).get(fn, 0)
        if val > 0:
            write_row(ws, row, [fn, "", "", "", "", "", val])
            row += 1
    row += 1

    # --- 明細 ---
    ws.cell(row=row, column=1, value="【明細】").font = BFONT
    row += 1
    write_header(ws, row, headers)
    row += 1

    for i, line in enumerate(data.get("lines", [])):
        write_row(ws, row, [
            line.get("floor", ""),
            line.get("molding_type", ""),
            line.get("cat", ""),
            line.get("type", ""),
            line.get("slope_side", ""),
            line.get("beam_name", ""),
            line.get("length_m", 0),
        ], is_alt=(i % 2 == 1))
        row += 1

    # 列幅調整
    widths = [6, 12, 10, 10, 8, 16, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    print(f"  廻り縁集計シート: {len(data.get('lines', []))}行")


def add_habaki_sheet(wb, data):
    """巾木集計シートを追加"""
    ws = wb.create_sheet("巾木集計")
    headers = ["階", "位置", "壁種別", "長さ(m)"]
    write_header(ws, 1, headers)

    row = 2
    # --- 階別集計 ---
    ws.cell(row=row, column=1, value="【階別集計】").font = BFONT
    row += 1
    sub_headers = ["階", "外周(片面)", "室内(両面)", "小計"]
    for ci, h in enumerate(sub_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = BFONT
        c.border = BORDER
    row += 1

    grand = 0
    for fn in ["1F", "2F"]:
        ft = data.get("floor_totals", {}).get(fn, {})
        if ft:
            ext = ft.get("ext", 0)
            intv = ft.get("int", 0)
            total = ft.get("total", ext + intv)
            write_row(ws, row, [fn, ext, intv, total])
            grand += total
            row += 1
    write_row(ws, row, ["合計", "", "", data.get("grand_total", grand)],
              is_bold=True, is_sub=True)
    row += 2

    # --- 明細 ---
    ws.cell(row=row, column=1, value="【明細】").font = BFONT
    row += 1
    write_header(ws, row, headers)
    row += 1

    for i, line in enumerate(data.get("lines", [])):
        pos = "外周" if line.get("type") == "exterior" else "室内"
        write_row(ws, row, [
            line.get("floor", ""),
            pos,
            line.get("cat", ""),
            line.get("length_m", 0),
        ], is_alt=(i % 2 == 1))
        row += 1

    widths = [6, 12, 10, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    print(f"  巾木集計シート: {len(data.get('lines', []))}行")


def add_gakubuchi_sheet(wb, data):
    """額縁集計シートを追加"""
    ws = wb.create_sheet("額縁集計")

    row = 1
    # --- 種別集計 ---
    headers_summary = ["部材種別", "合計長さ(m)"]
    write_header(ws, row, headers_summary)
    row += 1

    for kind in ["額縁", "額縁受け", "T-bar", "霧除け", "木口"]:
        val = data.get("type_totals", {}).get(kind, 0)
        if val > 0:
            write_row(ws, row, [kind, val])
            row += 1
    write_row(ws, row, ["合計", data.get("grand_total", 0)],
              is_bold=True, is_sub=True)
    row += 2

    # --- 建具別集計 ---
    ws.cell(row=row, column=1, value="【建具別集計】").font = BFONT
    row += 1
    fix_headers = ["建具名", "額縁タイプ", "IFC幅(mm)", "IFC高さ(mm)",
                   "額縁幅(mm)", "額縁高さ(mm)"]
    write_header(ws, row, fix_headers)
    row += 1
    for i, fix in enumerate(data.get("fixtures", [])):
        write_row(ws, row, [
            fix.get("label", ""),
            fix.get("ftype", ""),
            fix.get("w_mm", 0),
            fix.get("h_mm", 0),
            fix.get("frame_w", 0),
            fix.get("frame_h", 0),
        ], is_alt=(i % 2 == 1))
        row += 1
    row += 1

    # --- 部材明細 ---
    ws.cell(row=row, column=1, value="【部材明細】").font = BFONT
    row += 1
    detail_headers = ["部材種別", "建具名", "長さ(mm)"]
    write_header(ws, row, detail_headers)
    row += 1

    # 建具ごとにグループ化して集計
    by_fixture = defaultdict(list)
    for line in data.get("lines", []):
        by_fixture[line.get("fixture", "")].append(line)

    alt = False
    for fixture, lines in sorted(by_fixture.items()):
        for line in lines:
            write_row(ws, row, [
                line.get("kind", ""),
                fixture,
                line.get("length_mm", 0),
            ], is_alt=alt)
            row += 1
        alt = not alt

    widths = [12, 25, 12, 12, 12, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    print(f"  額縁集計シート: 建具{len(data.get('fixtures', []))}箇所, "
          f"部材{len(data.get('lines', []))}本")


def main():
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <buzai.xlsx> "
              f"[mawari_summary.json] [habaki_summary.json] [gakubuchi_summary.json]")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    # JSONパスは引数で明示するか、Excelと同じディレクトリから自動検出
    json_files = sys.argv[2:] if len(sys.argv) > 2 else []

    # 自動検出モード: Excelのbasenameからモデル名を推定
    if not json_files:
        out_dir = os.path.dirname(xlsx_path)
        basename = os.path.basename(xlsx_path)
        # buzai_MODEL.xlsx → MODEL
        model = basename.replace("buzai_", "").replace(".xlsx", "")
        candidates = [
            os.path.join(out_dir, f"{model}_summary.json"),         # mawari-buchi
            os.path.join(out_dir, f"habaki_{model}_summary.json"),   # habaki
            os.path.join(out_dir, f"gakubuchi_{model}_summary.json"),# gakubuchi
        ]
        json_files = [p for p in candidates if os.path.exists(p)]

    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel not found: {xlsx_path}")
        sys.exit(1)

    print(f"部材一覧Excel: {xlsx_path}")
    print(f"JSON集計ファイル: {len(json_files)}個")

    wb = load_workbook(xlsx_path)

    # 既存の追加シートがある場合は削除して再作成
    for sheet_name in ["廻り縁集計", "巾木集計", "額縁集計"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    added = 0
    for jpath in json_files:
        if not os.path.exists(jpath):
            print(f"  SKIP (not found): {jpath}")
            continue
        with open(jpath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        tool = data.get("tool", "")
        if tool == "廻り縁拾い":
            add_mawari_sheet(wb, data)
            added += 1
        elif tool == "巾木拾い":
            add_habaki_sheet(wb, data)
            added += 1
        elif tool == "額縁拾い":
            add_gakubuchi_sheet(wb, data)
            added += 1
        else:
            print(f"  SKIP (unknown tool: {tool}): {jpath}")

    wb.save(xlsx_path)
    print(f"\n{added}シートを追加 → {xlsx_path}")


if __name__ == "__main__":
    main()
